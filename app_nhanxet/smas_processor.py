# smas_processor.py - Xử lý file SMAS (Sổ đánh giá học sinh)
import os
import json
import random
import openpyxl
import xlrd

import sys
if getattr(sys, 'frozen', False):
    _APP_DIR = os.path.dirname(sys.executable)
else:
    _APP_DIR = os.path.dirname(os.path.abspath(__file__))

SETTINGS_FILE = os.path.join(_APP_DIR, "smas_settings.json")


def is_smas_file_xls(filepath):
    """Kiểm tra file .xls có phải file SMAS không.
    Dấu hiệu: nhiều sheet, mỗi sheet có 'BẢNG KẾT QUẢ ĐÁNH GIÁ MÔN' ở row 5."""
    try:
        wb = xlrd.open_workbook(filepath, on_demand=True)
        names = wb.sheet_names()
        if len(names) < 2:
            return False
        # Thử đọc sheet thứ 2 (sheet 0 thường bị lỗi)
        for idx in range(1, min(3, len(names))):
            try:
                ws = wb.sheet_by_index(idx)
                for r in range(min(8, ws.nrows)):
                    v = ws.cell_value(r, 0)
                    if v and isinstance(v, str) and "BẢNG KẾT QUẢ ĐÁNH GIÁ MÔN" in v.upper():
                        wb.release_resources()
                        return True
            except Exception:
                continue
        wb.release_resources()
    except Exception:
        pass
    return False


def is_smas_file_xlsx(filepath):
    """Kiểm tra file .xlsx có phải file SMAS không."""
    try:
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        names = wb.sheetnames
        if len(names) < 2:
            wb.close()
            return False
        for sn in names[1:3]:
            ws = wb[sn]
            for row in ws.iter_rows(min_row=1, max_row=8, max_col=1, values_only=True):
                v = row[0]
                if v and isinstance(v, str) and "BẢNG KẾT QUẢ ĐÁNH GIÁ MÔN" in v.upper():
                    wb.close()
                    return True
        wb.close()
    except Exception:
        pass
    return False


class SmasProcessor:
    """Xử lý file SMAS: đọc nhiều sheet môn học, điền nhận xét tự động."""

    def __init__(self):
        self.wb = None          # openpyxl workbook (để ghi)
        self.file_path = None
        self.file_ext = None
        self.school_name = ""
        self.class_name = ""
        self.year = ""
        self.grade_level = ""   # "Tiểu học", "THCS", "THPT"
        self.sheets_info = []   # [{name, subject, header_row, data_start, comment_cols, score_col, total_students}]

    def load_file(self, filepath):
        """Load file SMAS. Trả về dict thông tin tổng quan."""
        self.file_path = filepath
        self.file_ext = os.path.splitext(filepath)[1].lower()
        self.sheets_info = []

        if self.file_ext == ".xls":
            return self._load_xls(filepath)
        else:
            return self._load_xlsx(filepath)

    def _load_xls(self, filepath):
        """Load file .xls bằng xlrd, convert sang openpyxl GIỮ NGUYÊN FORMAT."""
        from openpyxl.styles import Font, Border, Side, Alignment, PatternFill
        from openpyxl.utils import get_column_letter

        # Try formatting_info=True (giữ format); fallback nếu file không hỗ trợ
        has_formatting = True
        try:
            xls_wb = xlrd.open_workbook(filepath, formatting_info=True, on_demand=False)
        except Exception:
            has_formatting = False
            xls_wb = xlrd.open_workbook(filepath, on_demand=True)
        sheet_names = xls_wb.sheet_names()

        self.wb = openpyxl.Workbook()
        default_ws = self.wb.active

        # Pre-build font map & format map from xlrd (chỉ khi có formatting_info)
        xf_list = xls_wb.xf_list if has_formatting else []
        font_list = xls_wb.font_list if has_formatting else []

        def _xlrd_colour_to_hex(colour_index, default="#000000"):
            """Convert xlrd colour index to hex."""
            if colour_index is None or colour_index < 8:
                return default
            colour_map = xls_wb.colour_map
            if colour_map and colour_index in colour_map:
                rgb = colour_map[colour_index]
                if rgb:
                    return f"#{rgb[0]:02X}{rgb[1]:02X}{rgb[2]:02X}"
            return default

        def _make_openpyxl_font(xlrd_font):
            """Convert xlrd font object to openpyxl Font."""
            try:
                color_hex = _xlrd_colour_to_hex(xlrd_font.colour_index, "000000")
                if color_hex.startswith("#"):
                    color_hex = color_hex[1:]
                return Font(
                    name=xlrd_font.name or "Arial",
                    size=xlrd_font.height / 20 if xlrd_font.height else 11,
                    bold=xlrd_font.bold,
                    italic=xlrd_font.italic,
                    underline="single" if xlrd_font.underlined else None,
                    color=color_hex
                )
            except Exception:
                return Font()

        def _make_border():
            """Create default thin border."""
            thin = Side(style="thin", color="000000")
            return Border(left=thin, right=thin, top=thin, bottom=thin)

        def _make_alignment(xf):
            """Convert xlrd XF alignment to openpyxl Alignment."""
            hor_map = {0: "general", 1: "left", 2: "center", 3: "right",
                       4: "fill", 5: "justify", 6: "centerContinuous"}
            ver_map = {0: "top", 1: "center", 2: "bottom"}
            try:
                return Alignment(
                    horizontal=hor_map.get(xf.alignment.hor_align, "general"),
                    vertical=ver_map.get(xf.alignment.vert_align, "bottom"),
                    wrap_text=bool(xf.alignment.text_wrap),
                )
            except Exception:
                return Alignment()

        first_valid = True
        for idx, sn in enumerate(sheet_names):
            try:
                xls_ws = xls_wb.sheet_by_index(idx)
            except Exception:
                continue

            if first_valid:
                ws = default_ws
                ws.title = sn
                first_valid = False
            else:
                ws = self.wb.create_sheet(title=sn)

            # Copy cell values + styles
            for r in range(xls_ws.nrows):
                for c in range(xls_ws.ncols):
                    val = xls_ws.cell_value(r, c)
                    ctype = xls_ws.cell_type(r, c)
                    if ctype == xlrd.XL_CELL_NUMBER and val == int(val):
                        val = int(val)

                    cell = ws.cell(row=r + 1, column=c + 1, value=val)

                    # Apply style from XF record (chỉ khi có formatting_info)
                    if has_formatting:
                        try:
                            xf_idx = xls_ws.cell_xf_index(r, c)
                            xf = xf_list[xf_idx]
                            font_idx = xf.font_index
                            if font_idx < len(font_list):
                                cell.font = _make_openpyxl_font(font_list[font_idx])
                            cell.alignment = _make_alignment(xf)
                            # Apply border if cell has border formatting
                            if xf.border_colour_indices and any(
                                b for b in xf.border_colour_indices if b):
                                cell.border = _make_border()
                        except Exception:
                            pass

            # Copy merged cells
            try:
                for crange in xls_ws.merged_cells:
                    rlo, rhi, clo, chi = crange
                    ws.merge_cells(
                        start_row=rlo + 1, start_column=clo + 1,
                        end_row=rhi, end_column=chi
                    )
            except Exception:
                pass

            # Copy column widths
            try:
                for col_idx in range(xls_ws.ncols):
                    col_letter = get_column_letter(col_idx + 1)
                    # xlrd colinfo_map: col_idx -> Colinfo object
                    if hasattr(xls_ws, 'colinfo_map') and col_idx in xls_ws.colinfo_map:
                        col_info = xls_ws.colinfo_map[col_idx]
                        # width in 1/256 of character width
                        width = col_info.width / 256
                        if width > 0:
                            ws.column_dimensions[col_letter].width = width
                    else:
                        ws.column_dimensions[col_letter].width = 12
            except Exception:
                pass

            # Copy row heights
            try:
                for row_idx in range(xls_ws.nrows):
                    if hasattr(xls_ws, 'rowinfo_map') and row_idx in xls_ws.rowinfo_map:
                        row_info = xls_ws.rowinfo_map[row_idx]
                        # height in twips (1/20 of a point)
                        height = row_info.height / 20
                        if height > 0:
                            ws.row_dimensions[row_idx + 1].height = height
            except Exception:
                pass

        xls_wb.release_resources()
        self._detect_all_sheets()

        # Nếu không có formatting_info → áp dụng format SMAS chuẩn
        if not has_formatting:
            for info in self.sheets_info:
                ws = self.wb[info["sheet_name"]]
                self._apply_smas_formatting(ws, info)

        return self._get_info()

    def _load_xlsx(self, filepath):
        """Load file .xlsx bằng openpyxl, GIỮ NGUYÊN FORMAT (không data_only)."""
        self.wb = openpyxl.load_workbook(filepath)
        self._detect_all_sheets()
        return self._get_info()

    def _apply_smas_formatting(self, ws, info):
        """Áp dụng format SMAS chuẩn cho sheet khi không copy được format gốc."""
        from openpyxl.styles import Font, Border, Side, Alignment
        from openpyxl.utils import get_column_letter

        max_col = ws.max_column or 8
        header_row = info["header_row"]
        data_start = info["data_start"]
        comment_cols = info.get("comment_cols", [])
        score_col = info.get("score_col")

        thin = Side(style="thin", color="000000")
        thin_border = Border(left=thin, right=thin, top=thin, bottom=thin)
        font_header = Font(name="Times New Roman", bold=True, size=11)
        font_title = Font(name="Times New Roman", bold=True, size=13)
        font_data = Font(name="Times New Roman", size=11)
        align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
        align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)

        # 1. Column widths and hidden columns
        for c in range(1, max_col + 1):
            letter = get_column_letter(c)
            col_title = str(ws.cell(header_row, c).value or "").strip().upper()
            
            if "STUDENTID" in col_title:
                ws.column_dimensions[letter].hidden = True
                
            if c == 1:
                ws.column_dimensions[letter].width = 5
            elif "HỌ VÀ TÊN" in col_title or "HỌ TÊN" in col_title:
                ws.column_dimensions[letter].width = 28
            elif c in comment_cols:
                ws.column_dimensions[letter].width = 45
            elif "MÃ" in col_title:
                ws.column_dimensions[letter].width = 14
            else:
                ws.column_dimensions[letter].width = 12

        # 2. Header area (trước bảng dữ liệu): merge, format, hide metadata
        for r in range(1, header_row):
            is_metadata = False
            first_val = str(ws.cell(r, 1).value or "").strip()
            
            # Row 1-3 usually contain UUIDs/numbers -> metadata
            if r <= 3 and (first_val.isdigit() or not first_val):
                is_metadata = True
                
            # Check for JSON or system tags
            for c in range(1, max_col + 1):
                val = str(ws.cell(r, c).value or "").strip()
                if val.startswith('{"Id":') or val.startswith('{"PointCode":'):
                    is_metadata = True
                    break
                if val in ["GHKII", "MHHD_NX_GK", "CHKII", "MHHD_NhanXet", "MHHD"]:
                    is_metadata = True
                    break
                    
            if is_metadata or not any(ws.cell(r, c).value for c in range(1, max_col + 1)):
                ws.row_dimensions[r].hidden = True
                continue

            # Dòng hiển thị (Tên trường, Bảng kết quả...)
            non_empty = []
            for c in range(1, max_col + 1):
                val = ws.cell(r, c).value
                if val is not None and str(val).strip():
                    non_empty.append((c, val))

            if 1 <= len(non_empty) <= 3:
                long_texts = [(c, v) for c, v in non_empty if isinstance(v, str) and len(v.strip()) > 5]
                if long_texts:
                    col_src, text = long_texts[0]
                    try:
                        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=max_col)
                    except Exception:
                        pass
                    cell = ws.cell(r, 1)
                    cell.value = text
                    if col_src != 1:
                        ws.cell(r, col_src).value = None
                    vu = text.upper()
                    if "BẢNG KẾT QUẢ" in vu or "ĐÁNH GIÁ" in vu:
                        cell.font = font_title
                    elif "TRƯỜNG" in vu:
                        cell.font = Font(name="Times New Roman", bold=True, size=12)
                    else:
                        cell.font = font_data
                    cell.alignment = Alignment(horizontal="center", vertical="center")

        # 3. Header row (STT, Họ tên, ...) + sub-header
        for hr in range(header_row, data_start):
            if not any(ws.cell(hr, c).value for c in range(1, max_col + 1)):
                ws.row_dimensions[hr].hidden = True
                continue
                
            ws.row_dimensions[hr].height = 30
            for c in range(1, max_col + 1):
                cell = ws.cell(hr, c)
                cell.font = font_header
                cell.alignment = align_center
                if not ws.column_dimensions[get_column_letter(c)].hidden:
                    cell.border = thin_border

        # 4. Data rows
        for r in range(data_start, ws.max_row + 1):
            stt = ws.cell(r, 1).value
            if stt is None or str(stt).strip() == "":
                continue
            try:
                int(float(str(stt)))
            except (ValueError, TypeError):
                continue
            for c in range(1, max_col + 1):
                cell = ws.cell(r, c)
                cell.font = font_data
                if not ws.column_dimensions[get_column_letter(c)].hidden:
                    cell.border = thin_border
                
                col_title = str(ws.cell(header_row, c).value or "").strip().upper()
                if "HỌ VÀ TÊN" in col_title or "HỌ TÊN" in col_title:
                    cell.alignment = Alignment(horizontal="left", vertical="center")
                elif c in comment_cols:
                    cell.alignment = align_left
                else:
                    cell.alignment = Alignment(horizontal="center", vertical="center")

    def _detect_all_sheets(self):
        """Phân tích cấu trúc từng sheet trong workbook."""
        self.sheets_info = []
        for sn in self.wb.sheetnames:
            ws = self.wb[sn]
            info = self._detect_sheet(ws, sn)
            if info:
                self.sheets_info.append(info)

    def _detect_sheet(self, ws, sheet_name):
        """Phân tích 1 sheet. Hỗ trợ cả SMAS lẫn file đánh giá generic."""
        subject = ""
        school = ""
        year = ""
        grade = ""
        class_name = ""

        # Bỏ qua sheet trống hoặc quá nhỏ
        if ws.max_row is None or ws.max_row < 3:
            return None

        # Scan header rows (1-10) trên NHIỀU cột
        for r in range(1, min(11, ws.max_row + 1)):
            for col in range(1, min(15, (ws.max_column or 1) + 1)):
                v = ws.cell(r, col).value
                if not v or not isinstance(v, str):
                    continue
                vu = v.upper().strip()

                if "TRƯỜNG" in vu and not school:
                    school = v.strip()
                    if not self.school_name:
                        self.school_name = school

                # Header SMAS: "BẢNG KẾT QUẢ ĐÁNH GIÁ MÔN..."
                if "BẢNG KẾT QUẢ ĐÁNH GIÁ MÔN" in vu and not subject:
                    text_after = v.upper().replace("BẢNG KẾT QUẢ ĐÁNH GIÁ MÔN", "").strip()
                    parts = text_after.split("-")
                    if parts:
                        # Lấy case gốc
                        try:
                            orig_after = v[v.upper().index("MÔN") + 3:].strip()
                            orig_parts = orig_after.split("-")
                            if orig_parts:
                                subject = orig_parts[0].strip()
                        except ValueError:
                            subject = parts[0].strip()
                    if len(parts) > 1:
                        for p in parts[1:]:
                            if "LỚP" in p.upper():
                                class_name = p.strip()

                # Header generic: "NỘI DUNG NHẬN XÉT", "ĐÁNH GIÁ ĐỊNH KỲ", etc.
                if not subject:
                    for kw in ["ĐÁNH GIÁ ĐỊNH KỲ", "ĐÁNH GIÁ NĂNG LỰC", "NỘI DUNG NHẬN XÉT",
                               "BẢNG ĐÁNH GIÁ", "ĐÁNH GIÁ HỌC SINH"]:
                        if kw in vu:
                            subject = sheet_name  # Dùng tên sheet làm subject
                            break

                if ("HỌC KỲ" in vu or "NĂM HỌC" in vu) and not year:
                    year = v.strip()
                    if not self.year:
                        self.year = year

                if ("LỚP" in vu or "KHỐI" in vu) and not class_name and len(v.strip()) < 30:
                    class_name = v.strip()

        # Detect grade from metadata row (col 7-8)
        for meta_col in [8, 7, 10]:
            meta_grade = ws.cell(1, meta_col).value
            if meta_grade and isinstance(meta_grade, str):
                mg = meta_grade.strip()
                if any(kw in mg.upper() for kw in ["TIỂU HỌC", "THCS", "THPT"]):
                    self.grade_level = mg
                    grade = mg
                    break

        # Fallback: nếu chưa có subject, dùng tên sheet (bỏ qua sheet tên lạ)
        if not subject:
            clean_name = sheet_name.strip()
            # Bỏ qua sheet có tên giống mã hệ thống
            if clean_name and clean_name != "Sheet1" and not clean_name.startswith("XXXX"):
                subject = clean_name
            else:
                # Thử quét thêm: có STT + Họ tên + dữ liệu → vẫn là sheet hợp lệ
                pass

        # Tìm header row (chứa "STT" hoặc "Họ và tên")
        header_row = None
        for r in range(1, min(20, ws.max_row + 1)):
            for c in range(1, min(10, (ws.max_column or 1) + 1)):
                v = ws.cell(r, c).value
                if v and isinstance(v, str):
                    vu = v.strip().upper()
                    if vu in ("STT", "SỐ THỨ TỰ", "TT"):
                        header_row = r
                        break
                    if "HỌ VÀ TÊN" in vu or "HỌ TÊN" in vu:
                        header_row = r
                        break
            if header_row:
                break

        if not header_row:
            return None

        # Nếu vẫn chưa có subject nhưng có header → dùng sheet_name
        if not subject:
            subject = sheet_name.strip() or f"Sheet"

        # Tìm tất cả cột "Nhận xét" / "Nội dung" trong header (quét nhiều dòng)
        comment_cols = []
        score_col = None
        name_col = None

        for c in range(1, (ws.max_column or 1) + 1):
            for check_r in range(max(1, header_row - 3), min(header_row + 2, ws.max_row + 1)):
                v = ws.cell(check_r, c).value
                if v and isinstance(v, str):
                    vu = v.upper().strip()
                    if "NHẬN XÉT" in vu or "NỘI DUNG" in vu:
                        if c not in comment_cols:
                            comment_cols.append(c)
                        break
                    if "HỌ VÀ TÊN" in vu or "HỌ TÊN" in vu:
                        name_col = c

        # Tìm cột điểm/mức
        for c in range(1, (ws.max_column or 1) + 1):
            for check_r in [header_row, header_row + 1]:
                if check_r > ws.max_row:
                    continue
                v = ws.cell(check_r, c).value
                if v and isinstance(v, str):
                    vu = v.upper().strip()
                    if any(kw in vu for kw in ["MỨC ĐẠT", "CUỐI HỌC KỲ", "CUỐI KỲ",
                                                "GIỮA HỌC KỲ", "ĐIỂM", "XẾP LOẠI", "MỨC"]):
                        score_col = c
                        if "CUỐI" in vu or "MỨC ĐẠT" in vu:
                            break

        # Fallback: nếu không tìm thấy cột nhận xét
        if not comment_cols:
            # Tìm cột cuối có header text
            for c in range((ws.max_column or 1), 0, -1):
                v = ws.cell(header_row, c).value
                if v and str(v).strip():
                    comment_cols = [c]
                    break

        if not comment_cols:
            return None

        # Fallback score_col
        if not score_col and comment_cols:
            score_col = comment_cols[-1] - 1
            if score_col < 1:
                score_col = 1

        # Data start row: sau header, bỏ qua sub-header
        data_start = header_row + 1
        # Kiểm tra 2 dòng tiếp theo xem có phải sub-header không
        for offset in range(3):
            check_row = header_row + 1 + offset
            if check_row > ws.max_row:
                break
            first_val = ws.cell(check_row, 1).value
            if first_val is not None:
                try:
                    int(float(str(first_val)))
                    data_start = check_row
                    break
                except (ValueError, TypeError):
                    continue

        # Đếm học sinh
        total = 0
        for r in range(data_start, ws.max_row + 1):
            stt = ws.cell(r, 1).value
            if stt is None or str(stt).strip() == "":
                continue
            try:
                int(float(str(stt)))
                total += 1
            except (ValueError, TypeError):
                continue

        if total == 0:
            return None

        if not self.class_name and class_name:
            self.class_name = class_name

        return {
            "sheet_name": sheet_name,
            "subject": subject,
            "header_row": header_row,
            "data_start": data_start,
            "comment_cols": comment_cols,
            "score_col": score_col,
            "total_students": total,
            "grade": grade,
        }

    def _get_info(self):
        """Trả về dict thông tin tổng quan."""
        subjects = [s["subject"] for s in self.sheets_info]
        total = sum(s["total_students"] for s in self.sheets_info)
        return {
            "school": self.school_name,
            "class": self.class_name,
            "year": self.year,
            "grade_level": self.grade_level,
            "total_sheets": len(self.sheets_info),
            "subjects": subjects,
            "total_students": total,
            "sheets_detail": self.sheets_info,
        }

    def get_sheet_names(self):
        """Trả về danh sách tên sheet hợp lệ."""
        return [s["sheet_name"] for s in self.sheets_info]

    def get_preview_data(self, sheet_name, max_rows=50):
        """Trả về (headers, rows) cho preview."""
        if not self.wb or sheet_name not in self.wb.sheetnames:
            return [], []

        ws = self.wb[sheet_name]
        info = None
        for s in self.sheets_info:
            if s["sheet_name"] == sheet_name:
                info = s
                break
        if not info:
            return [], []

        # Headers
        headers = []
        for c in range(1, ws.max_column + 1):
            v = ws.cell(info["header_row"], c).value
            headers.append(str(v) if v else f"Col{c}")

        # Rows
        rows = []
        for r in range(info["data_start"], min(info["data_start"] + max_rows, ws.max_row + 1)):
            row = []
            for c in range(1, ws.max_column + 1):
                v = ws.cell(r, c).value
                row.append(str(v) if v is not None else "")
            rows.append(row)

        return headers, rows

    def process_sheet(self, sheet_name, comment_bank=None, cap="tieu_hoc",
                      forced_subject=None, overwrite=False, manual_config=None):
        """Điền nhận xét cho 1 sheet cụ thể. Trả về stats dict."""
        if not self.wb:
            raise ValueError("Chưa load file!")

        info = None
        for s in self.sheets_info:
            if s["sheet_name"] == sheet_name:
                info = s
                break
        if not info:
            raise ValueError(f"Không tìm thấy sheet: {sheet_name}")

        ws = self.wb[sheet_name]
        filled = 0
        skipped = 0
        errors = 0
        details = []

        # Xác định cột nhận xét cần điền (cột nhận xét cuối cùng = cuối kỳ)
        if manual_config and manual_config.enabled:
            comment_col = manual_config.comment_col
            data_start = manual_config.row_start
            data_end = manual_config.row_end
            score_col = info["score_col"]
        else:
            comment_col = info["comment_cols"][-1] if info["comment_cols"] else None
            data_start = info["data_start"]
            data_end = ws.max_row
            score_col = info["score_col"]

        if not comment_col:
            raise ValueError("Không xác định được cột nhận xét!")

        subject_for_bank = forced_subject or info["subject"]

        for r in range(data_start, data_end + 1):
            stt = ws.cell(r, 1).value
            if stt is None or str(stt).strip() == "":
                continue
            try:
                int(float(str(stt)))
            except (ValueError, TypeError):
                continue

            existing = ws.cell(r, comment_col).value
            if existing and str(existing).strip() and not overwrite:
                skipped += 1
                continue

            # Lấy điểm/mức từ score_col
            score_val = ws.cell(r, score_col).value if score_col else None
            # Lấy tên HS (cột 4 = index 3 in 0-based, col 4 in 1-based)
            name = ws.cell(r, 4).value or ""
            name = str(name).strip()
            short_name = name.split()[-1] if name.split() else ""

            comment = None
            # Ưu tiên 1: Kho nhận xét theo môn
            if comment_bank and cap and subject_for_bank:
                level = self._score_to_level(score_val, cap)
                if level:
                    comment = comment_bank.get_random_comment(cap, "mon_hoc", subject_for_bank, level)
                    if not comment:
                        comment = self._bank_fallback(comment_bank, cap, subject_for_bank, level)

            # Ưu tiên 2: Sinh nhận xét mặc định
            if not comment:
                comment = self._generate_default_comment(score_val, short_name, subject_for_bank)

            if comment:
                ws.cell(r, comment_col, value=comment)
                filled += 1
                details.append(f"{name}: {score_val} → {comment[:40]}...")
            else:
                errors += 1
                details.append(f"{name}: không có điểm, bỏ qua")

        return {
            "sheet": sheet_name,
            "subject": info["subject"],
            "total": info["total_students"],
            "filled": filled,
            "skipped": skipped,
            "errors": errors,
            "details": details,
        }

    def process_all(self, comment_bank=None, cap="tieu_hoc",
                    forced_subject=None, overwrite=False):
        """Điền nhận xét cho TẤT CẢ sheet. Trả về tổng hợp stats."""
        all_stats = []
        total_filled = 0
        total_skipped = 0
        total_errors = 0

        for info in self.sheets_info:
            subj = forced_subject if forced_subject else None
            stats = self.process_sheet(
                info["sheet_name"], comment_bank, cap, subj, overwrite
            )
            all_stats.append(stats)
            total_filled += stats["filled"]
            total_skipped += stats["skipped"]
            total_errors += stats["errors"]

        return {
            "total_sheets": len(all_stats),
            "total_filled": total_filled,
            "total_skipped": total_skipped,
            "total_errors": total_errors,
            "per_sheet": all_stats,
        }

    def save_output(self, output_path):
        """Lưu workbook ra file."""
        if self.wb:
            self.wb.save(output_path)

    def _score_to_level(self, score_val, cap):
        """Chuyển mức T/H/C hoặc điểm số sang level cho kho nhận xét."""
        if score_val is None:
            return None

        val = str(score_val).strip().upper()

        # SMAS thường dùng T/H/C
        if val in ("T", "TỐT", "HOÀN THÀNH TỐT"):
            return "T"
        if val in ("H", "HOÀN THÀNH"):
            return "H" if cap == "tieu_hoc" else "K"
        if val in ("C", "CHƯA HOÀN THÀNH", "CHT"):
            return "C" if cap == "tieu_hoc" else "CD"
        if val in ("Đ", "ĐẠT", "DAT"):
            return "T" if cap == "tieu_hoc" else "D"
        if val in ("CĐ", "CHƯA ĐẠT"):
            return "C" if cap == "tieu_hoc" else "CD"

        # Thử parse số
        try:
            num = float(str(score_val).replace(",", "."))
            if cap == "tieu_hoc":
                if num >= 9: return "T"
                elif num >= 5: return "H"
                else: return "C"
            else:
                if num >= 8.5: return "XS"
                elif num >= 6.5: return "T"
                elif num >= 5: return "K"
                elif num >= 3.5: return "D"
                else: return "CD"
        except (ValueError, TypeError):
            return None

    def _bank_fallback(self, comment_bank, cap, subject, level):
        """Tìm nhận xét fallback từ kho (môn gần giống hoặc chung)."""
        subject_lower = subject.lower()
        all_subjects = comment_bank.get_subjects(cap, "mon_hoc") if hasattr(comment_bank, 'get_subjects') else []
        for s in all_subjects:
            if subject_lower in s.lower() or s.lower() in subject_lower:
                c = comment_bank.get_random_comment(cap, "mon_hoc", s, level)
                if c:
                    return c
        return None

    def _generate_default_comment(self, score_val, short_name, subject=""):
        """Sinh nhận xét mặc định khi kho không có."""
        if score_val is None:
            return None

        val = str(score_val).strip().upper()
        name = short_name or "em"

        templates = {
            "T": [
                f"Hoàn thành tốt môn {subject}. {name} tích cực, chăm chỉ trong học tập.",
                f"{name} đạt kết quả tốt, nắm vững kiến thức môn {subject}.",
                f"{name} có tinh thần học tập tích cực, hoàn thành tốt các yêu cầu của môn {subject}.",
            ],
            "H": [
                f"Hoàn thành môn {subject}. {name} cần cố gắng hơn để đạt kết quả tốt hơn.",
                f"{name} đạt yêu cầu môn {subject}, cần tích cực phát biểu xây dựng bài.",
                f"{name} hoàn thành các nội dung cơ bản của môn {subject}.",
            ],
            "C": [
                f"Chưa hoàn thành môn {subject}. {name} cần nỗ lực nhiều hơn trong học tập.",
                f"{name} cần ôn luyện thêm để nắm vững kiến thức cơ bản môn {subject}.",
            ],
        }

        level_key = None
        if val in ("T", "TỐT", "HOÀN THÀNH TỐT"):
            level_key = "T"
        elif val in ("H", "HOÀN THÀNH", "Đ", "ĐẠT"):
            level_key = "H"
        elif val in ("C", "CHT", "CHƯA HOÀN THÀNH", "CĐ", "CHƯA ĐẠT"):
            level_key = "C"
        else:
            try:
                num = float(str(score_val).replace(",", "."))
                if num >= 9:
                    level_key = "T"
                elif num >= 5:
                    level_key = "H"
                else:
                    level_key = "C"
            except (ValueError, TypeError):
                return None

        if level_key and level_key in templates:
            return random.choice(templates[level_key])
        return None
