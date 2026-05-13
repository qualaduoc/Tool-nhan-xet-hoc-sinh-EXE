# vnedu_subject_processor.py - Xử lý nhận xét sổ điểm chi tiết từng môn học
import os
import json
import random
import openpyxl
from grade_presets import GRADE_PRESETS, get_preset_as_settings

import sys
if getattr(sys, 'frozen', False):
    _APP_DIR = os.path.dirname(sys.executable)
else:
    _APP_DIR = os.path.dirname(os.path.abspath(__file__))
SETTINGS_FILE = os.path.join(_APP_DIR, "subject_comment_settings.json")

# Mặc định dùng preset THCS
DEFAULT_SETTINGS = get_preset_as_settings("thcs")


def load_subject_settings():
    if os.path.exists(SETTINGS_FILE):
        try:
            with open(SETTINGS_FILE, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            pass
    return DEFAULT_SETTINGS.copy()


def save_subject_settings(settings):
    with open(SETTINGS_FILE, "w", encoding="utf-8") as f:
        json.dump(settings, f, ensure_ascii=False, indent=2)


def is_subject_score_file(wb):
    """Kiểm tra xem workbook có phải là file sổ điểm chi tiết không.
    Trả về True nếu tìm thấy 'BẢNG ĐIỂM CHI TIẾT' trong header."""
    ws = wb.active
    for r in range(1, min(8, ws.max_row + 1)):
        for c in range(1, min(10, ws.max_column + 1)):
            v = ws.cell(r, c).value
            if v and isinstance(v, str) and "BẢNG ĐIỂM CHI TIẾT" in v.upper():
                return True
    return False


def is_subject_score_file_xls(sheet):
    """Kiểm tra file .xls có phải sổ điểm chi tiết không."""
    for r in range(0, min(7, sheet.nrows)):
        for c in range(0, min(10, sheet.ncols)):
            v = sheet.cell_value(r, c)
            if v and isinstance(v, str) and "BẢNG ĐIỂM CHI TIẾT" in v.upper():
                return True
    return False


class SubjectCommentProcessor:
    """Xử lý nhận xét tự động cho sổ điểm chi tiết từng môn VNEDU"""

    def __init__(self):
        self.wb = None
        self.ws = None
        self.file_path = None
        self.file_ext = None  # '.xls' or '.xlsx'
        self.subject_name = ""
        self.semester = ""
        self.year = ""
        self.class_name = ""
        self.school_name = ""
        self.score_col = None       # Cột điểm tham chiếu (trước cột Nhận xét)
        self.score_col_name = ""    # Tên cột điểm
        self.comment_col = None     # Cột Nhận xét
        self.score_type = "numeric" # "numeric" hoặc "text"
        self.header_row = None      # Hàng chứa header chính
        self.data_start_row = None  # Hàng bắt đầu dữ liệu HS
        self.total_students = 0
        self.settings = load_subject_settings()

    def load_file(self, filepath):
        """Load file và detect cấu trúc. Trả về dict thông tin."""
        self.file_path = filepath
        self.file_ext = os.path.splitext(filepath)[1].lower()

        if self.file_ext == ".xls":
            return self._load_xls(filepath)
        else:
            return self._load_xlsx(filepath)

    def _load_xlsx(self, filepath):
        """Load file .xlsx bằng openpyxl"""
        self.wb = openpyxl.load_workbook(filepath, data_only=True)
        self.ws = self.wb.active
        self._from_xls = False
        self._detect_structure_xlsx()
        return self._get_info()

    def _load_xls(self, filepath):
        """Load file .xls: đọc bằng xlrd, convert sang openpyxl để ghi"""
        import xlrd
        try:
            xls_wb = xlrd.open_workbook(filepath, formatting_info=True)
        except Exception:
            xls_wb = xlrd.open_workbook(filepath, formatting_info=False)
        xls_ws = xls_wb.sheet_by_index(0)

        # Convert sang openpyxl workbook
        self.wb = openpyxl.Workbook()
        self.ws = self.wb.active
        self.ws.title = xls_ws.name
        self._from_xls = True

        for r in range(xls_ws.nrows):
            for c in range(xls_ws.ncols):
                cell_value = xls_ws.cell_value(r, c)
                cell_type = xls_ws.cell_type(r, c)
                # Chuyển đổi số thực không cần thiết (1.0 → 1)
                if cell_type == xlrd.XL_CELL_NUMBER and cell_value == int(cell_value):
                    cell_value = int(cell_value)
                self.ws.cell(row=r + 1, column=c + 1, value=cell_value)

        # Copy merged cells
        try:
            for crange in xls_ws.merged_cells:
                rlo, rhi, clo, chi = crange
                self.ws.merge_cells(
                    start_row=rlo + 1, start_column=clo + 1,
                    end_row=rhi, end_column=chi
                )
        except Exception:
            pass

        xls_wb.release_resources()
        self._detect_structure_xlsx()
        return self._get_info()

    def _detect_structure_xlsx(self):
        """Auto-detect cấu trúc file: header, cột điểm, cột nhận xét"""
        ws = self.ws

        # 1. Tìm thông tin trường, môn, lớp từ header (hàng 1-6)
        for r in range(1, min(8, ws.max_row + 1)):
            for c in range(1, min(10, ws.max_column + 1)):
                v = ws.cell(r, c).value
                if not v or not isinstance(v, str):
                    continue
                v_upper = v.upper().strip()

                if "TRƯỜNG" in v_upper:
                    self.school_name = v.strip()

                if "BẢNG ĐIỂM CHI TIẾT" in v_upper:
                    # Trích xuất: "BẢNG ĐIỂM CHI TIẾT - MÔN TOÁN HỌC - HỌC KỲ 2 - NĂM HỌC 2025-2026"
                    parts = v.split("-")
                    for i, p in enumerate(parts):
                        p_strip = p.strip().upper()
                        if "MÔN" in p_strip:
                            self.subject_name = p.strip().replace("MÔN ", "").replace("môn ", "").strip()
                        if "HỌC KỲ" in p_strip or "HỌC KÌ" in p_strip:
                            self.semester = p.strip()
                        if "CUỐI KỲ" in p_strip or "CUỐI KÌ" in p_strip:
                            self.semester = p.strip()
                        if "NĂM HỌC" in p_strip:
                            self.year = p.strip()

                if "KHỐI" in v_upper or "LỚP" in v_upper:
                    self.class_name = v.strip()

        # 2. Tìm hàng header (chứa "STT" hoặc "Họ và tên")
        for r in range(1, min(12, ws.max_row + 1)):
            for c in range(1, min(10, ws.max_column + 1)):
                v = ws.cell(r, c).value
                if v and isinstance(v, str) and v.strip().upper() in ["STT", "SỐ THỨ TỰ"]:
                    self.header_row = r
                    break
            if self.header_row:
                break

        if not self.header_row:
            raise ValueError("Không tìm thấy hàng header (STT) trong file!")

        # 3. Tìm cột "Nhận xét"
        for c in range(1, ws.max_column + 1):
            v = ws.cell(self.header_row, c).value
            if v and isinstance(v, str) and "NHẬN XÉT" in v.upper().strip():
                self.comment_col = c
                break

        if not self.comment_col:
            raise ValueError("Không tìm thấy cột 'Nhận xét' trong file!")

        # 4. Cột điểm tham chiếu = cột ngay trước cột Nhận xét (có header)
        for c in range(self.comment_col - 1, 0, -1):
            v = ws.cell(self.header_row, c).value
            if v and str(v).strip():
                self.score_col = c
                self.score_col_name = str(v).strip()
                break

        if not self.score_col:
            # Fallback: dùng cột trước nhận xét
            self.score_col = self.comment_col - 1
            self.score_col_name = "Cột " + str(self.score_col)

        # 5. Xác định hàng dữ liệu bắt đầu
        self.data_start_row = self.header_row + 1
        # Kiểm tra nếu có sub-header (TX1, TX2...)
        first_val = ws.cell(self.data_start_row, 1).value
        if first_val and isinstance(first_val, str) and not first_val.strip().isdigit():
            self.data_start_row += 1

        # 6. Đếm số HS và detect loại điểm
        self.total_students = 0
        sample_scores = []
        for r in range(self.data_start_row, ws.max_row + 1):
            stt = ws.cell(r, 1).value
            if stt is None or str(stt).strip() == "":
                continue
            try:
                int(float(str(stt)))
            except (ValueError, TypeError):
                continue
            self.total_students += 1
            score_val = ws.cell(r, self.score_col).value
            if score_val is not None:
                sample_scores.append(score_val)

        # Detect loại: nếu phần lớn giá trị là chữ → text, ngược lại → numeric
        text_count = sum(1 for s in sample_scores if isinstance(s, str))
        if text_count > len(sample_scores) * 0.5:
            self.score_type = "text"
        else:
            self.score_type = "numeric"

    def _get_info(self):
        """Trả về dict thông tin file"""
        return {
            "school": self.school_name,
            "subject": self.subject_name,
            "semester": self.semester,
            "year": self.year,
            "class": self.class_name,
            "total_students": self.total_students,
            "score_col_name": self.score_col_name,
            "score_type": self.score_type,
            "comment_col": self.comment_col,
        }

    def process(self, overwrite=False, comment_bank=None, cap=None, forced_subject=None):
        """Điền nhận xét tự động. Trả về stats.
        comment_bank: CommentBank instance để lấy nhận xét theo môn cụ thể.
        cap: cấp học (tieu_hoc/thcs/thpt).
        forced_subject: tên môn GV đã chọn, ưu tiên lấy nhận xét từ comment_bank.
        """
        if not self.wb or not self.ws:
            raise ValueError("Chưa load file!")

        ws = self.ws
        settings = self.settings
        filled = 0
        skipped = 0
        errors = 0
        details = []

        # Tên môn để tra kho: ưu tiên forced_subject > self.subject_name
        subject_for_bank = forced_subject or self.subject_name

        for r in range(self.data_start_row, ws.max_row + 1):
            stt = ws.cell(r, 1).value
            if stt is None or str(stt).strip() == "":
                continue
            try:
                int(float(str(stt)))
            except (ValueError, TypeError):
                continue

            # Kiểm tra ô nhận xét đã có chưa
            existing = ws.cell(r, self.comment_col).value
            if existing and str(existing).strip() and not overwrite:
                skipped += 1
                continue

            # Lấy điểm
            score_val = ws.cell(r, self.score_col).value
            name_parts = []
            for c in [3, 4]:
                v = ws.cell(r, c).value
                if v and str(v).strip():
                    name_parts.append(str(v).strip())
            student_name = " ".join(name_parts)
            short_name = name_parts[-1] if name_parts else ""

            # === Ưu tiên 1: Lấy nhận xét từ kho theo môn cụ thể ===
            comment = None
            if comment_bank and cap and subject_for_bank:
                level = self._score_to_bank_level(score_val, cap)
                if level:
                    comment = comment_bank.get_random_comment(cap, "mon_hoc", subject_for_bank, level)
                    # Fallback: tìm môn gần giống trong kho
                    if not comment:
                        comment = self._bank_fallback(comment_bank, cap, subject_for_bank, level)

            # === Ưu tiên 2: Sinh nhận xét từ grade_presets (template) ===
            if not comment:
                comment = self._generate_comment(score_val, short_name)

            if comment:
                ws.cell(r, self.comment_col, value=comment)
                filled += 1
                details.append(f"{student_name}: {score_val} → {comment[:40]}...")
            else:
                errors += 1
                details.append(f"{student_name}: không có điểm, bỏ qua")

        return {
            "total": self.total_students,
            "filled": filled,
            "skipped": skipped,
            "errors": errors,
            "details": details
        }

    def _score_to_bank_level(self, score_val, cap):
        """Chuyển đổi điểm/mức từ VNEDU sang mã level của kho nhận xét.
        Tiểu học: T/H/C. THCS/THPT: XS/T/K/D/CD."""
        if score_val is None:
            return None

        # Nếu score_type là text (Đ/CĐ)
        if self.score_type == "text":
            val = str(score_val).strip().upper()
            if val in ("Đ", "ĐẠT", "DAT", "D"):
                return "T" if cap == "tieu_hoc" else "D"
            elif val in ("CĐ", "CĐẠ", "CHƯA ĐẠT", "CHUA DAT", "CD"):
                return "C" if cap == "tieu_hoc" else "CD"
            elif val in ("T", "TỐT", "TOT"):
                return "T"
            elif val in ("K", "KHÁ", "KHA"):
                return "K" if cap != "tieu_hoc" else "H"
            return "D" if cap != "tieu_hoc" else "H"

        # Nếu score_type là numeric (điểm số)
        try:
            score_num = float(str(score_val).replace(",", "."))
        except (ValueError, TypeError):
            return None

        if cap == "tieu_hoc":
            if score_num >= 9:
                return "T"
            elif score_num >= 5:
                return "H"
            else:
                return "C"
        else:  # thcs, thpt
            if score_num >= 9:
                return "T"
            elif score_num >= 7:
                return "K"
            elif score_num >= 5:
                return "D"
            else:
                return "CD"

    def _bank_fallback(self, comment_bank, cap, subject_name, level):
        """Tìm nhận xét từ môn gần giống trong kho"""
        subjects = comment_bank.data.get(cap, {}).get("mon_hoc", {})
        sn_lower = subject_name.lower()
        for key in subjects:
            if key.lower() in sn_lower or sn_lower in key.lower():
                comments = comment_bank.get_comments(cap, "mon_hoc", key, level)
                if comments:
                    return random.choice(comments)
        # Fallback mức chung
        muc_chung = comment_bank.data.get(cap, {}).get("muc_chung", {}).get(level, {})
        if isinstance(muc_chung, dict) and "nhan_xet" in muc_chung:
            pool = muc_chung["nhan_xet"]
            if pool:
                return random.choice(pool)
        return None

    def _generate_comment(self, score_val, short_name=""):
        """Sinh nhận xét dựa trên điểm và template"""
        if score_val is None:
            return None

        settings = self.settings

        if self.score_type == "text":
            # Đánh giá bằng chữ (Đ/CĐ)
            score_str = str(score_val).strip().upper()
            text_cfg = settings.get("text", DEFAULT_SETTINGS["text"])

            for level_key in ["dat", "chuadat"]:
                level = text_cfg.get(level_key, {})
                values = [v.upper() for v in level.get("values", [])]
                if score_str in values:
                    templates = level.get("templates", [])
                    if templates:
                        comment = random.choice(templates)
                        return comment.replace("[TÊN]", short_name)
            # Fallback
            return f"Đánh giá: {score_val}"

        else:
            # Đánh giá bằng điểm số
            try:
                score_num = float(str(score_val).replace(",", "."))
            except (ValueError, TypeError):
                return None

            num_cfg = settings.get("numeric", DEFAULT_SETTINGS["numeric"])

            # Sắp xếp theo min giảm dần
            levels = sorted(num_cfg.items(), key=lambda x: x[1].get("min", 0), reverse=True)
            for level_key, level in levels:
                if score_num >= level.get("min", 0):
                    templates = level.get("templates", [])
                    if templates:
                        comment = random.choice(templates)
                        return comment.replace("[TÊN]", short_name)
                    break

            return None

    def save_output(self, output_path):
        """Lưu file kết quả (.xlsx) với formatting chuyên nghiệp"""
        if self.wb:
            self._apply_output_formatting()
            self.wb.save(output_path)

    def _apply_output_formatting(self):
        """Apply formatting chuyên nghiệp cho file xuất ra"""
        from openpyxl.styles import Border, Side, Font, Alignment, PatternFill
        from openpyxl.utils import get_column_letter

        ws = self.ws
        if not ws or not self.header_row:
            return

        thin = Side(style='thin')
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        header_font = Font(name='Times New Roman', bold=True, size=11)
        data_font = Font(name='Times New Roman', size=11)
        title_font = Font(name='Times New Roman', bold=True, size=13)
        sub_title_font = Font(name='Times New Roman', bold=True, size=12)

        # 1. Tính chiều rộng cột dựa trên nội dung
        for c in range(1, ws.max_column + 1):
            max_len = 0
            for r in range(1, min(ws.max_row + 1, 60)):
                v = ws.cell(r, c).value
                if v is not None:
                    max_len = max(max_len, len(str(v)))
            col_letter = get_column_letter(c)
            calculated_width = min(max(max_len * 1.3 + 2, 6), 50)
            ws.column_dimensions[col_letter].width = calculated_width

        # Override chiều rộng cột đặc biệt
        ws.column_dimensions['A'].width = 5     # STT
        ws.column_dimensions['B'].width = 16    # Mã HS
        # Cột Họ tên (C+D)
        ws.column_dimensions[get_column_letter(3)].width = 18
        if ws.max_column >= 4:
            ws.column_dimensions[get_column_letter(4)].width = 10
        # Cột Nhận xét
        if self.comment_col:
            ws.column_dimensions[get_column_letter(self.comment_col)].width = 50

        # 2. Format tiêu đề (trước header)
        for r in range(1, self.header_row):
            for c in range(1, ws.max_column + 1):
                cell = ws.cell(r, c)
                if cell.value and isinstance(cell.value, str):
                    v_upper = cell.value.upper()
                    if 'BẢNG ĐIỂM' in v_upper:
                        cell.font = title_font
                    elif 'TRƯỜNG' in v_upper or 'KHỐI' in v_upper or 'LỚP' in v_upper:
                        cell.font = sub_title_font
                    else:
                        cell.font = Font(name='Times New Roman', size=11)
                    cell.alignment = Alignment(vertical='center')

        # 3. Format header (viền + bold + fill + center)
        sub_header_row = self.header_row + 1 if self.data_start_row > self.header_row + 1 else None
        for r_fmt in [self.header_row]:
            for c in range(1, ws.max_column + 1):
                cell = ws.cell(r_fmt, c)
                cell.border = border
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        if sub_header_row:
            for c in range(1, ws.max_column + 1):
                cell = ws.cell(sub_header_row, c)
                if cell.value:
                    cell.border = border
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        # 4. Format dữ liệu học sinh
        for r in range(self.data_start_row, ws.max_row + 1):
            stt = ws.cell(r, 1).value
            if stt is None and not any(ws.cell(r, cc).value for cc in range(1, min(5, ws.max_column + 1))):
                continue

            for c in range(1, ws.max_column + 1):
                cell = ws.cell(r, c)
                cell.border = border
                cell.font = data_font

                # Mã HS - hiển thị số nguyên (tránh 2.3E+09)
                if c == 2 and isinstance(cell.value, (int, float)):
                    cell.number_format = '0'

                # Canh giữa: STT, Mã HS, Điểm
                if c in [1, 2] or (c >= 5 and c != self.comment_col and c not in [3, 4]):
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                elif c == self.comment_col:
                    # Cột nhận xét: canh trái, wrap text
                    cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                else:
                    # Họ tên: canh trái
                    cell.alignment = Alignment(horizontal='left', vertical='center')

        # 5. Đặt chiều cao hàng hợp lý
        for r in range(self.data_start_row, ws.max_row + 1):
            ws.row_dimensions[r].height = 22
        ws.row_dimensions[self.header_row].height = 30
