# converter_processor.py - Engine chuyển đổi dữ liệu VNEDU ↔ CSDL Ngành
import openpyxl
import os
import re
from copy import copy
from difflib import SequenceMatcher
from excel_processor import load_excel_file

# === MAPPING CỘT: VNEDU_col → CSDL_col ===
VNEDU_TO_CSDL = {
    7: 8,    # Tiếng Việt - Mức
    8: 9,    # Tiếng Việt - Điểm
    9: 6,    # Toán - Mức
    10: 7,   # Toán - Điểm
    11: 10,  # Đạo đức - Mức
    12: 22,  # Âm nhạc - Mức
    13: 23,  # Mĩ thuật - Mức
    14: 25,  # GDTC - Mức
    15: 24,  # HĐTN - Mức
    16: 20,  # TH-CN (Công nghệ) - Mức
    17: 21,  # TH-CN (Công nghệ) - Điểm
    18: 15,  # TH-CN (Tin học) - Mức
    19: 16,  # TH-CN (Tin học) - Điểm
    20: 13,  # Tiếng Anh / Ngoại ngữ - Mức
    21: 14,  # Tiếng Anh / Ngoại ngữ - Điểm
    22: 17,  # Khoa học - Mức
    23: 18,  # Khoa học - Điểm
    24: 11,  # LS&ĐL - Mức
    25: 12,  # LS&ĐL - Điểm
    # Năng lực cốt lõi C26-C35 (cùng vị trí)
    26: 26, 27: 27, 28: 28, 29: 29, 30: 30,
    31: 31, 32: 32, 33: 33, 34: 34, 35: 35,
    # Phẩm chất C36-C39 (cùng vị trí)
    36: 36, 37: 37, 38: 38, 39: 39,
}

# Reverse mapping: CSDL → VNEDU
CSDL_TO_VNEDU = {v: k for k, v in VNEDU_TO_CSDL.items()}

# Tên mapping cho log
SUBJECT_NAMES = {
    7: "Tiếng Việt (Mức)", 8: "Tiếng Việt (Điểm)",
    9: "Toán (Mức)", 10: "Toán (Điểm)",
    11: "Đạo đức", 12: "Âm nhạc", 13: "Mĩ thuật",
    14: "GDTC", 15: "HĐTN",
    16: "CN-Công nghệ (Mức)", 17: "CN-Công nghệ (Điểm)",
    18: "CN-Tin học (Mức)", 19: "CN-Tin học (Điểm)",
    20: "Tiếng Anh (Mức)", 21: "Tiếng Anh (Điểm)",
    22: "Khoa học (Mức)", 23: "Khoa học (Điểm)",
    24: "LS&ĐL (Mức)", 25: "LS&ĐL (Điểm)",
    26: "NL-Tự chủ", 27: "NL-Giao tiếp", 28: "NL-GQVĐ",
    29: "NL-Ngôn ngữ", 30: "NL-Tính toán", 31: "NL-Khoa học",
    32: "NL-Công nghệ", 33: "NL-Tin học", 34: "NL-Thẩm mĩ", 35: "NL-Thể chất",
    36: "PC-Yêu nước", 37: "PC-Nhân ái", 38: "PC-Chăm chỉ", 39: "PC-Trung thực",
}

# Config hệ thống
VNEDU_DATA_START = 10
VNEDU_NAME_COL = 3
VNEDU_DOB_COL = 5

CSDL_DATA_START = 4
CSDL_NAME_COL = 4
CSDL_DOB_COL = 5


def normalize_name(name):
    """Chuẩn hóa tên để so sánh: lowercase, bỏ khoảng trắng thừa"""
    if not name:
        return ""
    s = str(name).strip().lower()
    s = re.sub(r'\s+', ' ', s)
    return s


def normalize_dob(dob):
    """Chuẩn hóa ngày sinh thành chuỗi dd/mm/yyyy"""
    if not dob:
        return ""
    s = str(dob).strip()
    # Xử lý datetime object
    if hasattr(dob, 'strftime'):
        return dob.strftime('%d/%m/%Y')
    return s


def detect_file_type(filepath):
    """Nhận diện loại file: 'vnedu' hoặc 'csdl'
    Dựa vào header đặc trưng của từng hệ thống.
    """
    wb = load_excel_file(filepath, data_only=True)
    ws = wb.active

    # VNEDU: R6 C1="STT", R6 C2="MÃ VNEDU"
    r6c1 = str(ws.cell(6, 1).value or "").strip().upper()
    r6c2 = str(ws.cell(6, 2).value or "").strip().upper()
    if "STT" in r6c1 and "VNEDU" in r6c2:
        wb.close()
        return "vnedu"

    # CSDL: R1 C1="STT", R1 C3 chứa "Mã định danh" hoặc R1 C2="Mã lớp"
    r1c1 = str(ws.cell(1, 1).value or "").strip().upper()
    r1c2 = str(ws.cell(1, 2).value or "").strip().upper()
    r1c3 = str(ws.cell(1, 3).value or "").strip().upper()
    if "STT" in r1c1 and ("MÃ LỚP" in r1c2 or "MÃ ĐỊNH DANH" in r1c3):
        wb.close()
        return "csdl"

    # Fallback: kiểm tra merged cells
    for mc in ws.merged_cells.ranges:
        val = str(ws.cell(mc.min_row, mc.min_col).value or "")
        if "VNEDU" in val.upper():
            wb.close()
            return "vnedu"
        if "MÃ ĐỊNH DANH" in val.upper():
            wb.close()
            return "csdl"

    wb.close()
    return None


def extract_students(ws, file_type):
    """Trích xuất danh sách học sinh từ worksheet.
    Returns: dict { normalized_key: { 'row': row_number, 'name': original_name, 'dob': dob } }
    """
    if file_type == "vnedu":
        start_row = VNEDU_DATA_START
        name_col = VNEDU_NAME_COL
        dob_col = VNEDU_DOB_COL
    else:
        start_row = CSDL_DATA_START
        name_col = CSDL_NAME_COL
        dob_col = CSDL_DOB_COL

    students = {}
    for r in range(start_row, ws.max_row + 1):
        name = ws.cell(r, name_col).value
        if not name or not str(name).strip():
            continue
        name_str = str(name).strip()
        dob = ws.cell(r, dob_col).value
        dob_str = normalize_dob(dob)
        key = normalize_name(name_str) + "|" + dob_str
        students[key] = {
            'row': r,
            'name': name_str,
            'dob': dob_str,
        }
    return students


class ConverterProcessor:
    def __init__(self):
        self.source_wb = None
        self.dest_wb = None
        self.source_type = None
        self.dest_type = None
        self.source_path = None
        self.dest_path = None
        self.logs = []

    def _log(self, msg):
        self.logs.append(msg)

    def load_source(self, filepath):
        """Tải file nguồn (đã có dữ liệu)"""
        self.source_type = detect_file_type(filepath)
        if not self.source_type:
            raise ValueError("Không nhận diện được loại file nguồn!\nFile phải là VNEDU hoặc CSDL Ngành.")
        self.source_path = filepath
        self.source_wb = load_excel_file(filepath, data_only=True)
        ws = self.source_wb.active
        students = extract_students(ws, self.source_type)
        label = "VNEDU" if self.source_type == "vnedu" else "CSDL Ngành"
        return {
            "type": label,
            "total_students": len(students),
            "filename": os.path.basename(filepath),
        }

    def load_dest(self, filepath):
        """Tải file đích (template / file cần điền dữ liệu vào)"""
        self.dest_type = detect_file_type(filepath)
        if not self.dest_type:
            raise ValueError("Không nhận diện được loại file đích!\nFile phải là VNEDU hoặc CSDL Ngành.")
        if self.source_type and self.dest_type == self.source_type:
            raise ValueError(
                f"File nguồn và file đích cùng loại ({self.dest_type})!\n"
                f"Vui lòng chọn 1 file VNEDU + 1 file CSDL Ngành."
            )
        self.dest_path = filepath
        # Load with formatting (không data_only) để giữ format
        self.dest_wb = load_excel_file(filepath)
        ws = self.dest_wb.active
        students = extract_students(ws, self.dest_type)
        label = "VNEDU" if self.dest_type == "vnedu" else "CSDL Ngành"
        return {
            "type": label,
            "total_students": len(students),
            "filename": os.path.basename(filepath),
        }

    def _copy_row_data(self, ws_src, ws_dst, src_row, dst_row, col_map):
        """Copy dữ liệu 1 HS theo mapping cột. Returns số ô đã điền."""
        count = 0
        for src_col, dst_col in col_map.items():
            val = ws_src.cell(src_row, src_col).value
            if val is not None and str(val).strip():
                ws_dst.cell(dst_row, dst_col, value=val)
                count += 1
        return count

    @staticmethod
    def _name_similarity(name1, name2):
        """Tính độ tương đồng giữa 2 tên (0.0 → 1.0)"""
        return SequenceMatcher(None, normalize_name(name1), normalize_name(name2)).ratio()

    def convert(self):
        """Chuyển đổi dữ liệu source → dest với 4 tầng match thông minh:
        T1: Tên chính xác + Ngày sinh khớp (100%)
        T2: Tên chính xác, bỏ ngày sinh (95%)
        T3: Tên tương tự ≥85% + Ngày sinh khớp (90%)
        T4: Tên tương tự ≥80%, bỏ ngày sinh (75%)
        """
        if not self.source_wb or not self.dest_wb:
            raise ValueError("Chưa tải đủ 2 file!")

        self.logs = []
        ws_src = self.source_wb.active
        ws_dst = self.dest_wb.active

        src_students = extract_students(ws_src, self.source_type)
        dst_students = extract_students(ws_dst, self.dest_type)

        self._log(f"File nguồn ({self.source_type.upper()}): {len(src_students)} học sinh")
        self._log(f"File đích ({self.dest_type.upper()}): {len(dst_students)} học sinh")
        self._log("")

        col_map = VNEDU_TO_CSDL if self.source_type == "vnedu" else CSDL_TO_VNEDU

        matched = 0
        not_found = 0
        cells_filled = 0
        matched_dst_rows = set()  # Tránh match trùng

        # Chuẩn bị lookup nhanh cho dst
        dst_by_name = {}       # normalized_name → [info]
        dst_by_dob = {}        # dob → [info]
        dst_list = list(dst_students.values())
        for di in dst_list:
            n = normalize_name(di['name'])
            dst_by_name.setdefault(n, []).append(di)
            if di['dob']:
                dst_by_dob.setdefault(di['dob'], []).append(di)

        for key, src_info in src_students.items():
            src_name = src_info['name']
            src_name_norm = normalize_name(src_name)
            src_dob = src_info['dob']
            src_row = src_info['row']
            match_info = None
            match_label = ""

            # === TẦNG 1: Tên chính xác + Ngày sinh khớp ===
            if key in dst_students and dst_students[key]['row'] not in matched_dst_rows:
                match_info = dst_students[key]
                match_label = "T1-chính xác"

            # === TẦNG 2: Tên chính xác, bỏ ngày sinh ===
            if not match_info and src_name_norm in dst_by_name:
                for di in dst_by_name[src_name_norm]:
                    if di['row'] not in matched_dst_rows:
                        match_info = di
                        match_label = "T2-tên khớp"
                        break

            # === TẦNG 3: Fuzzy tên ≥85% + Ngày sinh khớp ===
            if not match_info and src_dob and src_dob in dst_by_dob:
                best_score = 0
                best_di = None
                for di in dst_by_dob[src_dob]:
                    if di['row'] in matched_dst_rows:
                        continue
                    score = self._name_similarity(src_name, di['name'])
                    if score >= 0.85 and score > best_score:
                        best_score = score
                        best_di = di
                if best_di:
                    match_info = best_di
                    match_label = f"T3-fuzzy {best_score:.0%}+DOB"

            # === TẦNG 4: Fuzzy tên ≥80%, bỏ ngày sinh ===
            if not match_info:
                best_score = 0
                best_di = None
                for di in dst_list:
                    if di['row'] in matched_dst_rows:
                        continue
                    score = self._name_similarity(src_name, di['name'])
                    if score >= 0.80 and score > best_score:
                        best_score = score
                        best_di = di
                if best_di:
                    match_info = best_di
                    match_label = f"T4-fuzzy {best_score:.0%}"

            # === KẾT QUẢ ===
            if match_info:
                dst_row = match_info['row']
                count = self._copy_row_data(ws_src, ws_dst, src_row, dst_row, col_map)
                cells_filled += count
                matched += 1
                matched_dst_rows.add(dst_row)
                dst_name = match_info['name']
                if src_name != dst_name:
                    self._log(f"  ✅ {src_name} → {dst_name} (dòng {dst_row}, {count} ô) [{match_label}]")
                else:
                    self._log(f"  ✅ {src_name} → dòng {dst_row} ({count} ô) [{match_label}]")
            else:
                not_found += 1
                self._log(f"  ❌ Không tìm thấy: {src_name} ({src_dob})")

        direction = f"{self.source_type.upper()} → {self.dest_type.upper()}"
        self._log(f"\n--- KẾT QUẢ ({direction}) ---")
        self._log(f"Đã ghép: {matched}/{len(src_students)} học sinh")
        self._log(f"Không tìm thấy: {not_found} học sinh")
        self._log(f"Tổng ô đã điền: {cells_filled}")

        return {
            "matched": matched,
            "not_found": not_found,
            "cells_filled": cells_filled,
            "total_source": len(src_students),
            "total_dest": len(dst_students),
            "direction": direction,
            "details": self.logs,
        }

    def save_output(self, output_path):
        """Lưu file đích đã điền dữ liệu"""
        if self.dest_wb:
            self.dest_wb.save(output_path)

    def get_preview_data(self, which="source", max_rows=50):
        """Lấy dữ liệu preview cho Treeview"""
        if which == "source" and self.source_wb:
            ws = self.source_wb.active
            ft = self.source_type
        elif which == "dest" and self.dest_wb:
            ws = self.dest_wb.active
            ft = self.dest_type
        else:
            return None

        if ft == "vnedu":
            header_row = 9
            data_start = VNEDU_DATA_START
        else:
            header_row = 3
            data_start = CSDL_DATA_START

        # Headers
        headers = []
        for c in range(1, ws.max_column + 1):
            val = ws.cell(header_row, c).value
            headers.append(str(val) if val else f"Col{c}")

        # Data rows
        rows = []
        for r in range(data_start, min(data_start + max_rows, ws.max_row + 1)):
            row = []
            for c in range(1, ws.max_column + 1):
                val = ws.cell(r, c).value
                row.append(str(val).strip() if val is not None else "")
            if any(row):
                rows.append(row)

        return {"headers": headers, "rows": rows, "file_type": ft}
