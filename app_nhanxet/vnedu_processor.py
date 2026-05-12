# vnedu_processor.py - Engine xử lý file VNEDU
# Cấu trúc file VNEDU:
# - Row 1-5: Header (tên trường, bảng tổng hợp, năm học, lớp)
# - Row 6-8: Header nhóm (Môn học, Năng lực, Phẩm chất)
# - Row 9: Sub-header (Mức đạt được / Điểm KTĐK / tên năng lực, phẩm chất)
# - Row 10+: Dữ liệu học sinh
# Cột: A=STT, B=Mã VNEDU, C=Họ tên, D=(?), E=Ngày sinh, F=Nữ
# G(7)..Y(25): Môn học (xen kẽ Mức đạt được + Điểm KTĐK)
# Z(26)..AI(35): Năng lực cốt lõi
# AJ(36)..AN(40): Phẩm chất chủ yếu
# AO(41): Ghi chú

import openpyxl
import copy
import json
import os

# Ngưỡng điểm mặc định cho "Mức đạt được" môn học
DEFAULT_SCORE_RULES = {
    "T": {"min": 9, "max": 10, "label": "Hoàn thành tốt"},
    "H": {"min": 5, "max": 8.5, "label": "Hoàn thành"},
    "C": {"min": 0, "max": 4.5, "label": "Chưa hoàn thành"},
}

# Mapping cột: môn nào có Điểm KTĐK → cột điểm → cột mức
SUBJECT_SCORE_PAIRS = [
    # (col_muc, col_diem, ten_mon)
    (7, 8, "Tiếng Việt"),
    (9, 10, "Toán"),
    (18, 19, "Tin học"),  # Tin học và Công nghệ (Tin học)
    (20, 21, "Tiếng Anh"),
    (22, 23, "Khoa học"),
    (24, 25, "Lịch sử và Địa lí"),
    (16, 17, "Công nghệ"),  # Tin học và Công nghệ (Công nghệ)
]

# Cột chỉ có "Mức đạt được" (không có điểm) - đánh giá bằng nhận xét
SUBJECT_NOGRADE_COLS = [11, 12, 13, 14, 15]
# Đạo đức, Âm nhạc, Mĩ thuật, Giáo dục thể chất, Hoạt động trải nghiệm

# Cột Năng lực cốt lõi (26-35): T/Đ/C
NANGLUC_COLS = list(range(26, 36))
# Cột Phẩm chất chủ yếu (36-40): T/Đ/C
PHAMCHAT_COLS = list(range(36, 41))

import sys
if getattr(sys, 'frozen', False):
    _APP_DIR = os.path.dirname(sys.executable)
else:
    _APP_DIR = os.path.dirname(os.path.abspath(__file__))
SETTINGS_FILE = os.path.join(_APP_DIR, "vnedu_settings.json")


def load_settings():
    """Tải cài đặt ngưỡng điểm"""
    if os.path.exists(SETTINGS_FILE):
        try:
            with open(SETTINGS_FILE, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            pass
    return {
        "score_T_min": 9,
        "score_H_min": 5,
        "score_C_max": 4.5,
    }


def save_settings(settings):
    """Lưu cài đặt ngưỡng điểm"""
    with open(SETTINGS_FILE, "w", encoding="utf-8") as f:
        json.dump(settings, f, ensure_ascii=False, indent=2)


def score_to_level(score, settings=None):
    """
    Chuyển điểm số → mức đạt được (T/H/C)
    VD: 9 → T, 7 → H, 3 → C
    """
    if settings is None:
        settings = load_settings()

    try:
        score = float(score)
    except (ValueError, TypeError):
        return None

    t_min = float(settings.get("score_T_min", 9))
    h_min = float(settings.get("score_H_min", 5))

    if score >= t_min:
        return "T"
    elif score >= h_min:
        return "H"
    else:
        return "C"


class VneduProcessor:
    def __init__(self):
        self.wb = None
        self.ws = None
        self.file_path = None
        self.header_row = 9  # Row chứa "Mức đạt được" / "Điểm KTĐK"
        self.data_start_row = 10  # Row đầu tiên chứa data học sinh
        self.settings = load_settings()
        self.stats = {"total": 0, "filled": 0, "skipped": 0}

    def load_file(self, file_path):
        """Tải file VNEDU Excel"""
        self.file_path = file_path
        self.wb = openpyxl.load_workbook(file_path)
        self.ws = self.wb.active

        # Auto-detect header row và data start row
        self._detect_structure()
        return self._get_file_info()

    def _detect_structure(self):
        """Tự động nhận diện cấu trúc file"""
        for row in range(1, 15):
            for col in range(1, 50):
                val = self.ws.cell(row, col).value
                if val and "Mức đạt được" in str(val):
                    self.header_row = row
                    self.data_start_row = row + 1
                    return
        # Mặc định nếu không tìm thấy
        self.header_row = 9
        self.data_start_row = 10

    def _detect_score_pairs(self):
        """Tự động phát hiện cặp (Mức đạt được, Điểm KTĐK) từ header"""
        pairs = []
        hr = self.header_row
        max_col = self.ws.max_column or 50
        col = 1
        while col <= max_col:
            h_val = self.ws.cell(hr, col).value
            if h_val and "Mức đạt được" in str(h_val):
                # Kiểm tra cột kế tiếp có phải "Điểm KTĐK" không
                next_val = self.ws.cell(hr, col + 1).value if col + 1 <= max_col else None
                if next_val and "Điểm KTĐK" in str(next_val):
                    # Tìm tên môn từ header phía trên
                    subject = self._find_subject_name(col)
                    pairs.append((col, col + 1, subject))
                    col += 2
                    continue
            col += 1
        return pairs

    def _find_subject_name(self, col):
        """Tìm tên môn học từ các row header phía trên"""
        for row in range(self.header_row - 1, 5, -1):
            val = self.ws.cell(row, col).value
            if val and str(val).strip():
                return str(val).strip()
        return f"Môn (Cột {col})"

    def _get_file_info(self):
        """Lấy thông tin file"""
        info = {
            "school": "",
            "class": "",
            "year": "",
            "total_students": 0,
            "total_cols": self.ws.max_column,
        }
        # Đọc header
        for row in range(1, 6):
            val = self.ws.cell(row, 1).value
            if val:
                val = str(val).strip()
                if "TRƯỜNG" in val.upper():
                    info["school"] = val
                elif "Lớp" in val:
                    info["class"] = val
                elif "Năm học" in val:
                    info["year"] = val

        # Đếm học sinh
        for row in range(self.data_start_row, self.ws.max_row + 1):
            name = self.ws.cell(row, 3).value
            if name and str(name).strip():
                info["total_students"] += 1
            else:
                break
        self.stats["total"] = info["total_students"]
        return info

    def process(self):
        """
        Xử lý chính: điền "Mức đạt được" dựa trên "Điểm KTĐK"
        Trả về: stats dict
        """
        if not self.ws:
            return None

        # Phát hiện cặp cột tự động
        score_pairs = self._detect_score_pairs()
        self.stats = {"total": 0, "filled": 0, "skipped": 0, "details": []}

        for row in range(self.data_start_row, self.ws.max_row + 1):
            name = self.ws.cell(row, 3).value
            if not name or not str(name).strip():
                break
            self.stats["total"] += 1
            student_fills = 0

            for col_muc, col_diem, subject in score_pairs:
                diem = self.ws.cell(row, col_diem).value
                muc_hien_tai = self.ws.cell(row, col_muc).value

                if diem is not None and str(diem).strip():
                    level = score_to_level(diem, self.settings)
                    if level:
                        # Chỉ điền nếu ô trống hoặc ghi đè
                        if not muc_hien_tai or not str(muc_hien_tai).strip():
                            self.ws.cell(row, col_muc).value = level
                            student_fills += 1
                            self.stats["filled"] += 1
                        else:
                            self.stats["skipped"] += 1

            if student_fills > 0:
                self.stats["details"].append(
                    f"✅ {name}: Điền {student_fills} mức đạt được"
                )

        return self.stats

    def get_preview_data(self, max_rows=30):
        """Lấy dữ liệu preview cho UI"""
        if not self.ws:
            return None

        headers = []
        # Lấy header từ row 9
        for col in range(1, min(self.ws.max_column + 1, 42)):
            val = self.ws.cell(self.header_row, col).value
            if val:
                headers.append(str(val)[:15])
            else:
                # Thử lấy từ row 7-8
                for hr in [7, 8]:
                    hval = self.ws.cell(hr, col).value
                    if hval:
                        headers.append(str(hval)[:15])
                        break
                else:
                    headers.append(f"C{col}")

        rows = []
        for row in range(self.data_start_row, min(self.data_start_row + max_rows, self.ws.max_row + 1)):
            name = self.ws.cell(row, 3).value
            if not name:
                break
            row_data = []
            for col in range(1, min(self.ws.max_column + 1, 42)):
                v = self.ws.cell(row, col).value
                row_data.append(str(v) if v is not None else "")
            rows.append(row_data)

        return {"headers": headers, "rows": rows}

    def save_output(self, output_path):
        """Lưu file kết quả"""
        if self.wb:
            self.wb.save(output_path)
