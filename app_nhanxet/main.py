# main.py - Giao diện chính App ETA Insight
import customtkinter as ctk
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import os
from PIL import Image
from comment_data import CommentBank
from excel_processor import ExcelProcessor
from config_ui import ConfigWindow
from license_manager import check_license
from license_ui import ActivationScreen, LicenseInfoBar
from auto_updater import check_for_update_async, download_update_async, apply_update, get_current_version, get_update_log
from vnedu_processor import VneduProcessor, load_settings as vnedu_load_settings, save_settings as vnedu_save_settings, score_to_level
from converter_processor import ConverterProcessor
from vnedu_subject_processor import SubjectCommentProcessor, is_subject_score_file, load_subject_settings, save_subject_settings
from grade_presets import GRADE_PRESETS, get_preset_as_settings

ctk.set_appearance_mode("light")
ctk.set_default_color_theme("blue")

# Mapping hiển thị tiếng Việt
CAP_DISPLAY = {"tieu_hoc": "Tiểu Học", "thcs": "THCS", "thpt": "THPT"}
CAP_REVERSE = {"Tiểu Học": "tieu_hoc", "THCS": "thcs", "THPT": "thpt"}

ACCENT = "#E67E22"
ACCENT_HOVER = "#F39C12"
BG_MAIN = "#F5F0E8"
BG_CARD = "#FFFFFF"
BG_SIDEBAR = "#FFF0E0"
TEXT_DARK = "#2C3E50"
TEXT_MID = "#666666"
SUCCESS = "#27AE60"
DANGER = "#E74C3C"


# Đường dẫn tuyệt đối tới thư mục chứa EXE (hoặc main.py khi dev)
if getattr(sys, 'frozen', False):
    APP_DIR = os.path.dirname(sys.executable)
else:
    APP_DIR = os.path.dirname(os.path.abspath(__file__))


class MainApp(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.title(f"ETA Insight v{get_current_version()} — Đánh giá & Nhận xét Học sinh")
        self.geometry("1100x720")
        self.minsize(850, 580)
        self.configure(fg_color="#1A1A2E")

        # Set icon
        icon_path = os.path.join(APP_DIR, "..", "icon", "favicon.ico")
        if os.path.exists(icon_path):
            self.iconbitmap(icon_path)

        self.cb = CommentBank()
        self.processor = ExcelProcessor()
        self.vnedu = VneduProcessor()
        self.converter = ConverterProcessor()
        self.loaded_file = None
        self.vnedu_loaded_file = None
        self.config_win = None
        self.subject_proc = SubjectCommentProcessor()
        # Load settings đã lưu hoặc preset mặc định
        saved_subj = load_subject_settings()
        if saved_subj and saved_subj.get("numeric"):
            self.subject_proc.settings = saved_subj
        else:
            self.subject_proc.settings = get_preset_as_settings("thcs")
        self._vnedu_mode = "assessment"  # "assessment" hoặc "subject"

        # Thiết lập style hiện đại cho ttk.Treeview
        self.style = ttk.Style()
        self.style.theme_use("default")
        self.style.configure("Treeview",
                             background="#FFFFFF",
                             foreground="#34495E",
                             rowheight=28,
                             fieldbackground="#FFFFFF",
                             font=("Arial", 10),
                             borderwidth=0)
        self.style.configure("Treeview.Heading",
                             background="#F2F3F4",
                             foreground="#2C3E50",
                             font=("Arial", 10, "bold"),
                             borderwidth=0,
                             relief="flat")
        self.style.map("Treeview",
                       background=[("selected", "#D6EAF8")],
                       foreground=[("selected", "#154360")])
        self.style.layout("Treeview", [('Treeview.treearea', {'sticky': 'nswe'})])

        # Hiển thị Splash Screen trước
        self._show_splash()

    def _show_splash(self):
        """Splash Screen: hiện logo + tên app, check license trong background"""
        self.splash_frame = ctk.CTkFrame(self, fg_color="#1A1A2E")
        self.splash_frame.pack(fill="both", expand=True)

        # Logo mascot lớn
        mascot_path = os.path.join(APP_DIR, "mascot.png")
        if os.path.exists(mascot_path):
            try:
                mascot_img = ctk.CTkImage(Image.open(mascot_path), size=(120, 120))
                lbl = ctk.CTkLabel(self.splash_frame, image=mascot_img, text="")
                lbl.image = mascot_img  # giữ reference
                lbl.pack(pady=(100, 15))
            except Exception:
                ctk.CTkLabel(self.splash_frame, text="📝", font=("Arial", 72)).pack(pady=(100, 15))
        else:
            ctk.CTkLabel(self.splash_frame, text="📝", font=("Arial", 72)).pack(pady=(100, 15))

        # Tên app
        ctk.CTkLabel(self.splash_frame, text="ETA Insight",
                     font=("Arial", 32, "bold"), text_color="#E67E22").pack()
        ctk.CTkLabel(self.splash_frame, text=f"v{get_current_version()}", 
                     font=("Arial", 14), text_color="#F39C12").pack(pady=(5, 8))
        ctk.CTkLabel(self.splash_frame, text="Đánh giá & Nhận xét Học sinh Tự động",
                     font=("Arial", 13), text_color="#AAB7C4").pack()

        # Loading indicator
        self._splash_status = ctk.CTkLabel(self.splash_frame, text="Đang khởi động...",
                     font=("Arial", 11), text_color="#888888")
        self._splash_status.pack(pady=(25, 5))
        self._splash_progress = ctk.CTkProgressBar(self.splash_frame, width=250,
                     height=4, fg_color="#2D2D44", progress_color="#E67E22")
        self._splash_progress.pack()
        self._splash_progress.configure(mode="indeterminate")
        self._splash_progress.start()

        # Footer splash
        ctk.CTkLabel(self.splash_frame, text="© 2026 Nguyễn Thành Được — Cộng đồng ETA",
                     font=("Arial", 10), text_color="#555555").pack(side="bottom", pady=30)

        # Check license trong BACKGROUND THREAD (tránh block UI)
        self._license_result = None
        import threading
        def _bg_check():
            self._license_result = check_license()
        t = threading.Thread(target=_bg_check, daemon=True)
        t.start()
        # Poll kết quả mỗi 100ms
        self.after(100, self._poll_license)

    def _poll_license(self):
        """Poll kết quả check_license từ background thread"""
        if self._license_result is None:
            # Vẫn đang chờ → poll tiếp
            self.after(100, self._poll_license)
            return
        # Có kết quả → chuyển tiếp
        self._splash_progress.stop()
        self.splash_frame.destroy()
        activated, msg, expiry = self._license_result
        if activated:
            self.configure(fg_color=BG_MAIN)
            self._build_ui()
        else:
            self._show_activation()

    def _show_activation(self):
        """Hiển thị màn hình kích hoạt bản quyền"""
        # Xóa mọi widget cũ
        for w in self.winfo_children():
            w.destroy()
        self.configure(fg_color="#1A1A2E")
        screen = ActivationScreen(self, on_success=self._on_activated)
        screen.pack(fill="both", expand=True)

    def _on_activated(self):
        """Callback khi kích hoạt thành công → load app chính"""
        for w in self.winfo_children():
            w.destroy()
        self.configure(fg_color=BG_MAIN)
        self._build_ui()

    def _build_ui(self):
        # === LICENSE BAR ===
        LicenseInfoBar(self, on_deactivate=self._show_activation).pack(fill="x")

        # === TOP BAR ===
        topbar = ctk.CTkFrame(self, height=60, fg_color=ACCENT, corner_radius=0)
        topbar.pack(fill="x")
        topbar.pack_propagate(False)

        # Logo mascot
        mascot_path = os.path.join(APP_DIR, "mascot.png")
        if os.path.exists(mascot_path):
            try:
                mascot_img = ctk.CTkImage(Image.open(mascot_path), size=(42, 42))
                ctk.CTkLabel(topbar, image=mascot_img, text="").pack(side="left", padx=(15,5))
            except Exception:
                pass

        # Tên app
        title_frame = ctk.CTkFrame(topbar, fg_color="transparent")
        title_frame.pack(side="left", padx=(0,10))
        ctk.CTkLabel(title_frame, text=f"ETA Insight",
                     font=("Arial", 18, "bold"), text_color="white").pack(anchor="w")
        ctk.CTkLabel(title_frame, text="Đánh giá & Nhận xét Học sinh Tự động",
                     font=("Arial", 10), text_color="#FFE0B2").pack(anchor="w")

        # Nút cập nhật (ẩn, chỉ hiện khi có bản mới)
        self._update_info = None
        self.update_btn = ctk.CTkButton(topbar, text="", width=0, height=30,
                                         fg_color="transparent", hover_color="#D35400",
                                         font=("Arial", 11, "bold"), text_color="#FFFFFF",
                                         corner_radius=15, command=self._on_update_click)
        # Chưa pack — chỉ pack khi phát hiện bản mới

        # Version badge bên phải
        ctk.CTkLabel(topbar, text=f"v{get_current_version()}",
                     font=("Arial", 11, "bold"), text_color="#FFE0B2",
                     fg_color="#C0392B", corner_radius=10, width=50, height=22).pack(side="right", padx=(0,15))

        # Kiểm tra cập nhật ngầm
        self._blink_state = True
        self._update_info = None
        check_for_update_async(self._on_update_check_done)

        # === NAVBAR ===
        navbar = ctk.CTkFrame(self, height=40, fg_color="#34495E", corner_radius=0)
        navbar.pack(fill="x")
        navbar.pack_propagate(False)
        self._active_page = "csdl"
        self.nav_btn_csdl = ctk.CTkButton(navbar, text="📋 CSDL Ngành", width=160, height=34,
                                            fg_color=ACCENT, hover_color=ACCENT_HOVER,
                                            font=("Arial", 13, "bold"), corner_radius=6,
                                            command=lambda: self._switch_page("csdl"))
        self.nav_btn_csdl.pack(side="left", padx=(15,5), pady=3)
        self.nav_btn_vnedu = ctk.CTkButton(navbar, text="🌐 VNEDU", width=140, height=34,
                                            fg_color="transparent", hover_color="#4A6FA5",
                                            font=("Arial", 13, "bold"), corner_radius=6,
                                            text_color="#AAB7C4",
                                            command=lambda: self._switch_page("vnedu"))
        self.nav_btn_vnedu.pack(side="left", padx=5, pady=3)
        self.nav_btn_convert = ctk.CTkButton(navbar, text="🔄 Chuyển Đổi", width=160, height=34,
                                              fg_color="transparent", hover_color="#4A6FA5",
                                              font=("Arial", 13, "bold"), corner_radius=6,
                                              text_color="#AAB7C4",
                                              command=lambda: self._switch_page("convert"))
        self.nav_btn_convert.pack(side="left", padx=5, pady=3)

        # === CONTENT CONTAINER ===
        self.content = ctk.CTkFrame(self, fg_color=BG_MAIN, corner_radius=0)
        self.content.pack(fill="both", expand=True)

        # Tạo 3 trang
        self.page_csdl = ctk.CTkFrame(self.content, fg_color=BG_MAIN, corner_radius=0)
        self.page_vnedu = ctk.CTkFrame(self.content, fg_color=BG_MAIN, corner_radius=0)
        self.page_convert = ctk.CTkFrame(self.content, fg_color=BG_MAIN, corner_radius=0)
        self._build_page_csdl(self.page_csdl)
        self._build_page_vnedu(self.page_vnedu)
        self._build_page_convert(self.page_convert)

        # Hiện trang mặc định
        self.page_csdl.pack(fill="both", expand=True)

        # === BOTTOM BAR ===
        bottom = ctk.CTkFrame(self, height=36, fg_color="#1A252F", corner_radius=0)
        bottom.pack(fill="x", side="bottom")
        bottom.pack_propagate(False)
        ctk.CTkLabel(bottom, text=f"ETA Insight v{get_current_version()} © 2026 | Nguyễn Thành Được — Cộng đồng ETA",
                     font=("Arial", 10, "bold"), text_color="#AEB6BF").pack(side="left", padx=15)
        ctk.CTkLabel(bottom, text="📞 0904059866  ✉ nguyenthanhduocathy@gmail.com",
                     font=("Arial", 10), text_color="#7F8C8D").pack(side="right", padx=15)

    def _switch_page(self, page_name):
        """Chuyển giữa 3 trang: CSDL Ngành / VNEDU / Chuyển đổi"""
        if page_name == self._active_page:
            return
        self._active_page = page_name

        self.page_csdl.pack_forget()
        self.page_vnedu.pack_forget()
        self.page_convert.pack_forget()

        # Reset tất cả nút về trạng thái inactive
        inactive = {"fg_color": "transparent", "text_color": "#AAB7C4"}
        self.nav_btn_csdl.configure(**inactive)
        self.nav_btn_vnedu.configure(**inactive)
        self.nav_btn_convert.configure(**inactive)

        if page_name == "csdl":
            self.nav_btn_csdl.configure(fg_color=ACCENT, text_color="white")
            self.page_csdl.pack(fill="both", expand=True)
        elif page_name == "vnedu":
            self.nav_btn_vnedu.configure(fg_color="#3498DB", text_color="white")
            self.page_vnedu.pack(fill="both", expand=True)
        elif page_name == "convert":
            self.nav_btn_convert.configure(fg_color="#8E44AD", text_color="white")
            self.page_convert.pack(fill="both", expand=True)

    def _build_page_csdl(self, parent):
        """Xây dựng trang CSDL Ngành (nội dung hiện tại)"""

        # === MAIN LAYOUT: PanedWindow cho co giãn tối ưu ===
        paned = tk.PanedWindow(parent, orient="horizontal", sashwidth=6,
                               bg="#D5C9B8", sashrelief="flat")
        paned.pack(fill="both", expand=True, padx=10, pady=(10,5))

        # === LEFT: Upload & Process (scrollable) ===
        left_wrapper = ctk.CTkFrame(paned, fg_color=BG_CARD, corner_radius=12,
                                     border_width=1, border_color="#E0D5C5")
        paned.add(left_wrapper, minsize=380, stretch="always")

        left = ctk.CTkScrollableFrame(left_wrapper, fg_color=BG_CARD, corner_radius=0)
        left.pack(fill="both", expand=True)

        # Section 1: Upload
        s1 = ctk.CTkFrame(left, fg_color="transparent")
        s1.pack(fill="x", padx=20, pady=(20,10))
        ctk.CTkLabel(s1, text="1. TẢI FILE EXCEL", font=("Arial", 13, "bold"),
                     text_color=TEXT_DARK).pack(anchor="w")
        ctk.CTkLabel(s1, text="Chọn file .xlsx từ máy tính (file đánh giá học sinh).",
                     font=("Arial", 11), text_color=TEXT_MID).pack(anchor="w", pady=(2,8))

        btn_row = ctk.CTkFrame(s1, fg_color="transparent")
        btn_row.pack(fill="x")
        ctk.CTkButton(btn_row, text="📂 Chọn File Excel...", fg_color="#FFFFFF",
                       text_color=ACCENT, border_width=1, border_color=ACCENT,
                       hover_color="#FDEBD0", font=("Arial", 12, "bold"),
                       height=36, width=160, command=self._open_file).pack(side="left", padx=(0,10))
        ctk.CTkButton(btn_row, text="⚙ Cấu Hình Lời Nhận Xét", fg_color="#2C3E50",
                       hover_color="#34495E", font=("Arial", 11),
                       height=36, width=160, command=self._open_config).pack(side="left")

        self.file_label = ctk.CTkLabel(s1, text="Chưa chọn file", font=("Arial", 11, "italic"),
                                       text_color=TEXT_MID)
        self.file_label.pack(anchor="w", pady=(8,0))

        # Separator
        ctk.CTkFrame(left, height=1, fg_color="#EAECEE").pack(fill="x", padx=20, pady=5)

        # Section 2: File info
        s2 = ctk.CTkFrame(left, fg_color="transparent")
        s2.pack(fill="x", padx=20, pady=5)
        ctk.CTkLabel(s2, text="2. THÔNG TIN FILE", font=("Arial", 13, "bold"),
                     text_color=TEXT_DARK).pack(anchor="w")

        self.info_frame = ctk.CTkFrame(s2, fg_color="#FFFFFF", border_width=1, border_color="#E0E0E0", corner_radius=6)
        self.info_frame.pack(fill="x", pady=(5,0))
        self.info_label = ctk.CTkLabel(self.info_frame, text="Tải file để xem thông tin...",
                                       font=("Arial", 11), text_color=TEXT_MID, wraplength=350,
                                       justify="left")
        self.info_label.pack(padx=12, pady=10, anchor="w")

        # Separator
        ctk.CTkFrame(left, height=1, fg_color="#EAECEE").pack(fill="x", padx=20, pady=10)

        # Section 3: Settings
        s3 = ctk.CTkFrame(left, fg_color="transparent")
        s3.pack(fill="x", padx=20, pady=5)
        ctk.CTkLabel(s3, text="3. CẤU HÌNH XỬ LÝ", font=("Arial", 13, "bold"),
                     text_color=TEXT_DARK).pack(anchor="w")

        opt_frame = ctk.CTkFrame(s3, fg_color="#FFFFFF", border_width=1, border_color="#E0E0E0", corner_radius=6)
        opt_frame.pack(fill="x", pady=8)

        ctk.CTkLabel(opt_frame, text="Chọn Cấp Học:", font=("Arial", 11, "bold"), text_color=TEXT_DARK).pack(anchor="w", padx=12, pady=(10,2))
        self.cap_display_var = ctk.StringVar(value="Tiểu Học")
        cap_menu = ctk.CTkSegmentedButton(opt_frame, values=["Tiểu Học", "THCS", "THPT"],
                                           variable=self.cap_display_var,
                                           font=("Arial", 11),
                                           selected_color=ACCENT, selected_hover_color=ACCENT_HOVER)
        cap_menu.pack(padx=12, pady=(0,8), fill="x")

        self.overwrite_var = ctk.BooleanVar(value=False)
        ctk.CTkCheckBox(opt_frame, text="Ghi đè ô nhận xét đã có sẵn",
                        variable=self.overwrite_var, font=("Arial", 11),
                        text_color=TEXT_DARK, fg_color=ACCENT,
                        hover_color=ACCENT_HOVER).pack(anchor="w", padx=12, pady=(0,10))

        # Separator
        ctk.CTkFrame(left, height=1, fg_color="#EAECEE").pack(fill="x", padx=20, pady=5)

        # Section 4: Actions
        s4 = ctk.CTkFrame(left, fg_color="transparent")
        s4.pack(fill="x", padx=20, pady=(10,15))
        ctk.CTkLabel(s4, text="4. THỰC THI", font=("Arial", 13, "bold"),
                     text_color=TEXT_DARK).pack(anchor="w", pady=(0,8))

        action_frame = ctk.CTkFrame(s4, fg_color="transparent")
        action_frame.pack(fill="x")

        self.run_btn = ctk.CTkButton(action_frame, text="🚀 ĐIỀN NHẬN XÉT TỰ ĐỘNG",
                                      fg_color=ACCENT, hover_color=ACCENT_HOVER,
                                      font=("Arial", 12, "bold"), height=40,
                                      command=self._run_process, state="disabled")
        self.run_btn.pack(fill="x", pady=(0,8))

        self.export_btn = ctk.CTkButton(action_frame, text="💾 XUẤT FILE KẾT QUẢ",
                                         fg_color=SUCCESS, hover_color="#219A52",
                                         font=("Arial", 12, "bold"), height=40,
                                         command=self._export_file, state="disabled")
        self.export_btn.pack(fill="x")

        # === RIGHT: Preview & Log ===
        right = ctk.CTkFrame(paned, fg_color=BG_CARD, corner_radius=12,
                             border_width=1, border_color="#E0D5C5")
        paned.add(right, minsize=350, stretch="always")

        # Title bar cho preview
        preview_header = ctk.CTkFrame(right, fg_color="#2C3E50", corner_radius=8, height=40)
        preview_header.pack(fill="x", padx=12, pady=(12,0))
        preview_header.pack_propagate(False)
        ctk.CTkLabel(preview_header, text="📋 XEM TRƯỚC DỮ LIỆU",
                     font=("Arial", 13, "bold"), text_color="white").pack(side="left", padx=15)
        self.preview_stats = ctk.CTkLabel(preview_header, text="",
                                          font=("Arial", 10), text_color="#82E0AA")
        self.preview_stats.pack(side="right", padx=15)

        # Sheet tabs container
        self.sheet_tabs_frame = ctk.CTkFrame(right, fg_color="#ECF0F1", corner_radius=0, height=35)
        self.sheet_tabs_frame.pack(fill="x", padx=12)
        self.sheet_tabs_frame.pack_propagate(False)
        self._current_sheets = []
        self._current_sheet_idx = 0

        # Preview
        self.preview_frame = ctk.CTkFrame(right, fg_color="#FFFFFF", corner_radius=0)
        self.preview_frame.pack(fill="both", expand=True, padx=12, pady=(0,5))

        # Placeholder khi chưa có file
        self.preview_placeholder = ctk.CTkFrame(self.preview_frame, fg_color="transparent")
        self.preview_placeholder.pack(fill="both", expand=True, pady=40)
        ctk.CTkLabel(self.preview_placeholder, text="📂",
                     font=("Arial", 42), text_color="#BDC3C7").pack()
        ctk.CTkLabel(self.preview_placeholder, text="Tải file Excel để xem trước dữ liệu",
                     font=("Arial", 13), text_color="#7F8C8D").pack(pady=(10,0))
        ctk.CTkLabel(self.preview_placeholder, text="Hỗ trợ file .xlsx đánh giá học sinh",
                     font=("Arial", 11), text_color="#BDC3C7").pack()

        # Biến để giữ widget Treeview
        self.preview_tree = None

        # Log section
        log_header = ctk.CTkFrame(right, fg_color="#2C3E50", corner_radius=8, height=32)
        log_header.pack(fill="x", padx=12, pady=(5,0))
        log_header.pack_propagate(False)
        ctk.CTkLabel(log_header, text="📝 NHẬT KÝ XỬ LÝ",
                     font=("Arial", 11, "bold"), text_color="white").pack(side="left", padx=15)

        self.log_box = ctk.CTkTextbox(right, height=120, fg_color="#1A252F", text_color="#2ECC71",
                                      font=("Consolas", 11), corner_radius=0,
                                      border_width=1, border_color="#34495E")
        self.log_box.pack(fill="x", padx=12, pady=(0,12))
        self._log("Ứng dụng sẵn sàng. Hãy tải file Excel để bắt đầu!")

    def _build_page_vnedu(self, parent):
        """Xây dựng trang VNEDU"""
        paned = tk.PanedWindow(parent, orient="horizontal", sashwidth=6,
                               bg="#D5C9B8", sashrelief="flat")
        paned.pack(fill="both", expand=True, padx=10, pady=(10,5))

        # === LEFT: Upload & Cấu hình ===
        left_wrapper = ctk.CTkFrame(paned, fg_color=BG_CARD, corner_radius=12,
                                     border_width=1, border_color="#E0D5C5")
        paned.add(left_wrapper, minsize=380, stretch="always")
        left = ctk.CTkScrollableFrame(left_wrapper, fg_color=BG_CARD, corner_radius=0)
        left.pack(fill="both", expand=True)

        # Section 1: Upload file VNEDU
        s1 = ctk.CTkFrame(left, fg_color="transparent")
        s1.pack(fill="x", padx=20, pady=(20,10))
        ctk.CTkLabel(s1, text="1. TẢI FILE VNEDU", font=("Arial", 13, "bold"),
                     text_color=TEXT_DARK).pack(anchor="w")
        ctk.CTkLabel(s1, text="File tổng hợp đánh giá hoặc Sổ điểm chi tiết (.xls/.xlsx)",
                     font=("Arial", 11), text_color=TEXT_MID).pack(anchor="w", pady=(2,8))

        btn_row = ctk.CTkFrame(s1, fg_color="transparent")
        btn_row.pack(fill="x")
        ctk.CTkButton(btn_row, text="📂 Chọn File VNEDU...", fg_color="#FFFFFF",
                       text_color="#3498DB", border_width=1, border_color="#3498DB",
                       hover_color="#EBF5FB", font=("Arial", 12, "bold"),
                       height=36, width=160, command=self._vnedu_open_file).pack(side="left", padx=(0,10))
        ctk.CTkButton(btn_row, text="⚙ Cấu Hình Lời Nhận Xét", fg_color="#2C3E50",
                       hover_color="#34495E", font=("Arial", 11),
                       height=36, width=160, command=self._open_subject_config).pack(side="left")

        # Badge hiện loại file
        self.vnedu_mode_badge = ctk.CTkLabel(btn_row, text="", font=("Arial", 10, "bold"),
                                              corner_radius=6, width=0, height=22)
        self.vnedu_mode_badge.pack(side="left", padx=(10,0))

        self.vnedu_file_label = ctk.CTkLabel(s1, text="Chưa chọn file", font=("Arial", 11, "italic"),
                                              text_color=TEXT_MID)
        self.vnedu_file_label.pack(anchor="w", pady=(8,0))

        ctk.CTkFrame(left, height=1, fg_color="#EAECEE").pack(fill="x", padx=20, pady=5)

        # Section 2: Thông tin
        s2 = ctk.CTkFrame(left, fg_color="transparent")
        s2.pack(fill="x", padx=20, pady=5)
        ctk.CTkLabel(s2, text="2. THÔNG TIN FILE", font=("Arial", 13, "bold"),
                     text_color=TEXT_DARK).pack(anchor="w")
        self.vnedu_info_frame = ctk.CTkFrame(s2, fg_color="#FFFFFF", border_width=1, border_color="#E0E0E0", corner_radius=6)
        self.vnedu_info_frame.pack(fill="x", pady=(5,0))
        self.vnedu_info_label = ctk.CTkLabel(self.vnedu_info_frame, text="Vui lòng tải file để xem thông tin lớp...",
                                              font=("Arial", 11), text_color=TEXT_MID,
                                              wraplength=350, justify="left")
        self.vnedu_info_label.pack(padx=12, pady=10, anchor="w")

        ctk.CTkFrame(left, height=1, fg_color="#EAECEE").pack(fill="x", padx=20, pady=10)

        # Container cho Section 3 (chứa cả 3A và 3B, chỉ hiện 1)
        self.vnedu_s3_container = ctk.CTkFrame(left, fg_color="transparent")
        self.vnedu_s3_container.pack(fill="x", padx=20, pady=5)

        # === Section 3A: Cấu hình ngưỡng điểm (mode assessment) ===
        self.vnedu_s3a = ctk.CTkFrame(self.vnedu_s3_container, fg_color="transparent")
        self.vnedu_s3a.pack(fill="x")
        ctk.CTkLabel(self.vnedu_s3a, text="3. CẤU HÌNH NGƯỠNG ĐIỂM", font=("Arial", 13, "bold"),
                     text_color=TEXT_DARK).pack(anchor="w")

        score_frame = ctk.CTkFrame(self.vnedu_s3a, fg_color="#FFFFFF", border_width=1, border_color="#E0E0E0", corner_radius=6)
        score_frame.pack(fill="x", pady=8)

        settings = vnedu_load_settings()

        r1 = ctk.CTkFrame(score_frame, fg_color="transparent")
        r1.pack(fill="x", padx=12, pady=(10,3))
        ctk.CTkLabel(r1, text="T (Hoàn thành tốt):", font=("Arial", 11, "bold"),
                     text_color="#27AE60", width=120, anchor="w").pack(side="left")
        ctk.CTkLabel(r1, text="từ", font=("Arial", 11), text_color=TEXT_MID).pack(side="left", padx=(5,5))
        self.vnedu_t_min = ctk.CTkEntry(r1, width=45, height=26, font=("Arial", 12), justify="center", border_width=1)
        self.vnedu_t_min.insert(0, str(settings.get("score_T_min", 9)))
        self.vnedu_t_min.pack(side="left")

        r2 = ctk.CTkFrame(score_frame, fg_color="transparent")
        r2.pack(fill="x", padx=12, pady=3)
        ctk.CTkLabel(r2, text="H (Hoàn thành):", font=("Arial", 11, "bold"),
                     text_color="#E67E22", width=120, anchor="w").pack(side="left")
        ctk.CTkLabel(r2, text="từ", font=("Arial", 11), text_color=TEXT_MID).pack(side="left", padx=(5,5))
        self.vnedu_h_min = ctk.CTkEntry(r2, width=45, height=26, font=("Arial", 12), justify="center", border_width=1)
        self.vnedu_h_min.insert(0, str(settings.get("score_H_min", 5)))
        self.vnedu_h_min.pack(side="left")

        r3 = ctk.CTkFrame(score_frame, fg_color="transparent")
        r3.pack(fill="x", padx=12, pady=(3,8))
        ctk.CTkLabel(r3, text="C (Chưa HT):", font=("Arial", 11, "bold"),
                     text_color="#E74C3C", width=120, anchor="w").pack(side="left")
        ctk.CTkLabel(r3, text="< ngưỡng H", font=("Arial", 11), text_color=TEXT_MID).pack(side="left", padx=5)

        ctk.CTkButton(score_frame, text="Lưu cấu hình", fg_color="#F2F4F4", text_color="#2C3E50",
                       hover_color="#E5E8E8", height=28, width=100, font=("Arial", 11),
                       command=self._vnedu_save_settings).pack(padx=12, pady=(0,10), anchor="e")

        # === Section 3B: Cấu hình môn học (ẩn ban đầu, nút đã nằm ở Section 1) ===
        self.vnedu_s3b = ctk.CTkFrame(self.vnedu_s3_container, fg_color="transparent")
        # Chỉ cần state, nút đã ở Section 1 đồng bộ với CSDL Ngành
        self.subj_level_widgets = []
        self.subj_text_widgets = {}
        self._subj_grade_key = "thcs"
        self._subject_config_win = None

        ctk.CTkFrame(left, height=1, fg_color="#EAECEE").pack(fill="x", padx=20, pady=5)

        # Section 4: Actions
        s4 = ctk.CTkFrame(left, fg_color="transparent")
        s4.pack(fill="x", padx=20, pady=(10,15))
        ctk.CTkLabel(s4, text="4. THỰC THI", font=("Arial", 13, "bold"),
                     text_color=TEXT_DARK).pack(anchor="w", pady=(0,8))

        action_frame = ctk.CTkFrame(s4, fg_color="transparent")
        action_frame.pack(fill="x")
        
        self.vnedu_run_btn = ctk.CTkButton(action_frame, text="🚀 ĐIỀN MỨC ĐẠT ĐƯỢC",
                                            fg_color="#2C3E50", hover_color="#34495E",
                                            font=("Arial", 12, "bold"), height=40,
                                            command=self._vnedu_run, state="disabled")
        self.vnedu_run_btn.pack(fill="x", pady=(0,8))

        self.vnedu_subj_run_btn = ctk.CTkButton(action_frame, text="📝 NHẬN XÉT MÔN HỌC",
                                                 fg_color=ACCENT, hover_color=ACCENT_HOVER,
                                                 font=("Arial", 12, "bold"), height=40,
                                                 command=self._vnedu_subject_run, state="disabled")

        self.vnedu_export_btn = ctk.CTkButton(action_frame, text="💾 XUẤT FILE KẾT QUẢ",
                                               fg_color=SUCCESS, hover_color="#219A52",
                                               font=("Arial", 12, "bold"), height=40,
                                               command=self._vnedu_export, state="disabled")
        self.vnedu_export_btn.pack(fill="x")

        # === RIGHT: Preview & Log ===
        right = ctk.CTkFrame(paned, fg_color=BG_CARD, corner_radius=12,
                             border_width=1, border_color="#E0D5C5")
        paned.add(right, minsize=350, stretch="always")

        # Preview header
        preview_header = ctk.CTkFrame(right, fg_color="#1A5276", corner_radius=8, height=40)
        preview_header.pack(fill="x", padx=12, pady=(12,0))
        preview_header.pack_propagate(False)
        ctk.CTkLabel(preview_header, text="📋 XEM TRƯỚC DỮ LIỆU VNEDU",
                     font=("Arial", 13, "bold"), text_color="white").pack(side="left", padx=15)
        self.vnedu_preview_stats = ctk.CTkLabel(preview_header, text="",
                                                 font=("Arial", 10), text_color="#82E0AA")
        self.vnedu_preview_stats.pack(side="right", padx=15)

        # Preview
        self.vnedu_preview_frame = ctk.CTkFrame(right, fg_color="#FFFFFF", corner_radius=0)
        self.vnedu_preview_frame.pack(fill="both", expand=True, padx=12, pady=(5,5))

        # Placeholder
        self.vnedu_ph = ctk.CTkFrame(self.vnedu_preview_frame, fg_color="transparent")
        self.vnedu_ph.pack(fill="both", expand=True, pady=40)
        ctk.CTkLabel(self.vnedu_ph, text="🌐", font=("Arial", 42), text_color="#BDC3C7").pack()
        ctk.CTkLabel(self.vnedu_ph, text="Tải file VNEDU để xem trước dữ liệu",
                     font=("Arial", 13), text_color="#7F8C8D").pack(pady=(10,0))
        
        self.vnedu_tree = None

        # Log VNEDU
        log_header = ctk.CTkFrame(right, fg_color="#1A5276", corner_radius=8, height=32)
        log_header.pack(fill="x", padx=12, pady=(5,0))
        log_header.pack_propagate(False)
        ctk.CTkLabel(log_header, text="📝 NHẬT KÝ XỬ LÝ VNEDU",
                     font=("Arial", 11, "bold"), text_color="white").pack(side="left", padx=15)

        self.vnedu_log_box = ctk.CTkTextbox(right, height=120, fg_color="#1A252F",
                                             text_color="#5DADE2", font=("Consolas", 11),
                                             corner_radius=0, border_width=1, border_color="#34495E")
        self.vnedu_log_box.pack(fill="x", padx=12, pady=(0,12))
        self._vnedu_log("Chọn file VNEDU (.xlsx) để bắt đầu!")

    # =======================================
    # VNEDU METHODS
    # =======================================
    def _vnedu_log(self, text):
        self.vnedu_log_box.insert("end", f"→ {text}\n")
        self.vnedu_log_box.see("end")

    def _vnedu_save_settings(self):
        """Lưu cấu hình ngưỡng điểm VNEDU"""
        try:
            t_min = float(self.vnedu_t_min.get())
            h_min = float(self.vnedu_h_min.get())
            if h_min >= t_min:
                messagebox.showwarning("Lỗi", "Ngưỡng H phải nhỏ hơn ngưỡng T!")
                return
            settings = {"score_T_min": t_min, "score_H_min": h_min}
            vnedu_save_settings(settings)
            self.vnedu.settings = settings
            self._vnedu_log(f"💾 Đã lưu: T ≥ {t_min}, H ≥ {h_min}, C < {h_min}")
            messagebox.showinfo("Đã lưu", f"Cấu hình ngưỡng điểm đã được lưu!\n\nT (Hoàn thành tốt): ≥ {t_min}\nH (Hoàn thành): ≥ {h_min}\nC (Chưa hoàn thành): < {h_min}")
        except ValueError:
            messagebox.showerror("Lỗi", "Vui lòng nhập số hợp lệ!")

    def _vnedu_open_file(self):
        """Mở file VNEDU — auto-detect loại file"""
        filepath = filedialog.askopenfilename(
            title="Chọn file Excel VNEDU",
            filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")]
        )
        if not filepath:
            return

        self.vnedu_loaded_file = filepath
        filename = os.path.basename(filepath)
        self.vnedu_file_label.configure(text=f"✅ {filename}", text_color=SUCCESS)
        self._vnedu_log(f"Đã tải: {filename}")

        # Auto-detect loại file
        try:
            ext = os.path.splitext(filepath)[1].lower()
            is_subject = False

            if ext == ".xls":
                import xlrd
                xls_wb = xlrd.open_workbook(filepath, formatting_info=False)
                from vnedu_subject_processor import is_subject_score_file_xls
                is_subject = is_subject_score_file_xls(xls_wb.sheet_by_index(0))
                xls_wb.release_resources()
            else:
                import openpyxl
                test_wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
                is_subject = is_subject_score_file(test_wb)
                test_wb.close()

            if is_subject:
                self._vnedu_switch_mode("subject")
                # Load bằng SubjectCommentProcessor
                info = self.subject_proc.load_file(filepath)
                info_text = f"📚 SỔ ĐIỂM CHI TIẾT\n"
                info_text += f"Trường: {info.get('school', 'N/A')}\n"
                info_text += f"Môn: {info.get('subject', 'N/A')}\n"
                info_text += f"HK: {info.get('semester', '')} - {info.get('year', '')}\n"
                info_text += f"Lớp: {info.get('class', '')}\n"
                info_text += f"Số HS: {info.get('total_students', 0)}\n"
                info_text += f"Cột điểm: {info.get('score_col_name', '?')}\n"
                info_text += f"Loại: {'Chữ (Đ/CĐ)' if info.get('score_type') == 'text' else 'Điểm số'}"
                self.vnedu_info_label.configure(text=info_text, text_color=TEXT_DARK)
                self.vnedu_preview_stats.configure(text=f"📊 {info.get('total_students', 0)} HS — {info.get('subject', '')}")
                self._vnedu_log(f"✅ Nhận diện: Sổ điểm [{info.get('subject', '')}] — {info.get('total_students', 0)} HS")
                self._vnedu_log(f"   Cột điểm tham chiếu: {info.get('score_col_name', '?')}")
                self.vnedu_subj_run_btn.configure(state="normal")
                # Auto-detect cấp học
                self._auto_detect_grade(info)
                # Preview dùng subject_proc
                self._vnedu_show_preview_subject()
            else:
                self._vnedu_switch_mode("assessment")
                info = self.vnedu.load_file(filepath)
                info_text = f"Trường: {info.get('school', 'N/A')}\n"
                info_text += f"{info.get('class', '')}\n"
                info_text += f"{info.get('year', '')}\n"
                info_text += f"Tổng số học sinh: {info.get('total_students', 0)}"
                self.vnedu_info_label.configure(text=info_text, text_color=TEXT_DARK)
                self.vnedu_preview_stats.configure(text=f"📊 {info.get('total_students', 0)} học sinh")
                self._vnedu_show_preview()
                self.vnedu_run_btn.configure(state="normal")
                self._vnedu_log(f"Sẵn sàng! {info.get('total_students', 0)} học sinh.")

        except Exception as e:
            self._vnedu_log(f"LỖI: {str(e)}")
            messagebox.showerror("Lỗi", f"Không thể đọc file:\n{str(e)}")

    def _vnedu_switch_mode(self, mode):
        """Chuyển đổi UI giữa mode assessment và subject"""
        self._vnedu_mode = mode
        # Reset buttons
        self.vnedu_export_btn.configure(state="disabled")

        if mode == "subject":
            self.vnedu_mode_badge.configure(text=" 📚 SỔ ĐIỂM ", fg_color="#E67E22", text_color="white")
            # Ẩn section 3A, hiện 3B
            self.vnedu_s3a.pack_forget()
            self.vnedu_s3b.pack(fill="x")
            # Ẩn nút assessment, hiện nút subject
            self.vnedu_run_btn.pack_forget()
            self.vnedu_subj_run_btn.pack(fill="x", pady=(0,8))
            self.vnedu_export_btn.pack_forget()
            self.vnedu_export_btn.pack(fill="x")
            self.vnedu_run_btn.configure(state="disabled")
        else:
            self.vnedu_mode_badge.configure(text=" 📝 TỔNG HỢP ", fg_color="#3498DB", text_color="white")
            # Ẩn section 3B, hiện 3A
            self.vnedu_s3b.pack_forget()
            self.vnedu_s3a.pack(fill="x")
            # Ẩn nút subject, hiện nút assessment
            self.vnedu_subj_run_btn.pack_forget()
            self.vnedu_run_btn.pack(fill="x", pady=(0,8))
            self.vnedu_export_btn.pack_forget()
            self.vnedu_export_btn.pack(fill="x")
            self.vnedu_subj_run_btn.configure(state="disabled")

    def _vnedu_show_preview_subject(self):
        """Hiển thị preview cho sổ điểm chi tiết"""
        for w in self.vnedu_preview_frame.winfo_children():
            w.destroy()

        ws = self.subject_proc.ws
        if not ws:
            return

        # Lấy header
        headers = []
        for c in range(1, min(20, ws.max_column + 1)):
            v = ws.cell(self.subject_proc.header_row, c).value
            headers.append(str(v) if v else f"Col{c}")

        tree = ttk.Treeview(self.vnedu_preview_frame, columns=list(range(len(headers))),
                            show="headings", height=15)
        for i, h in enumerate(headers):
            tree.heading(i, text=h[:15])
            tree.column(i, width=80, minwidth=50)

        # Dữ liệu
        for r in range(self.subject_proc.data_start_row, min(self.subject_proc.data_start_row + 100, ws.max_row + 1)):
            values = []
            for c in range(1, min(20, ws.max_column + 1)):
                v = ws.cell(r, c).value
                values.append(str(v) if v is not None else "")
            tree.insert("", "end", values=values)

        scrollbar_y = ttk.Scrollbar(self.vnedu_preview_frame, orient="vertical", command=tree.yview)
        scrollbar_x = ttk.Scrollbar(self.vnedu_preview_frame, orient="horizontal", command=tree.xview)
        tree.configure(yscrollcommand=scrollbar_y.set, xscrollcommand=scrollbar_x.set)

        tree.pack(side="left", fill="both", expand=True)
        scrollbar_y.pack(side="right", fill="y")
        scrollbar_x.pack(side="bottom", fill="x")

        self.vnedu_tree = tree

    def _vnedu_subject_run(self):
        """Chạy nhận xét tự động cho sổ điểm chi tiết"""
        if not self.subject_proc.wb:
            messagebox.showwarning("Chưa có file", "Vui lòng tải file sổ điểm trước!")
            return

        # Cập nhật settings từ UI
        self._subj_collect_settings()

        self._vnedu_log("Bắt đầu nhận xét môn học...")
        try:
            stats = self.subject_proc.process()
            self._vnedu_log(f"✅ Hoàn tất! {stats['total']} học sinh")
            self._vnedu_log(f"   Đã nhận xét: {stats['filled']} ô")
            self._vnedu_log(f"   Bỏ qua (đã có): {stats['skipped']} ô")
            if stats['errors']:
                self._vnedu_log(f"   Lỗi/bỏ qua: {stats['errors']} ô")

            for d in stats.get("details", [])[:8]:
                self._vnedu_log(f"   {d}")

            self.vnedu_export_btn.configure(state="normal")
            self._vnedu_show_preview_subject()

            messagebox.showinfo("Thành công",
                f"Đã nhận xét tự động!\n\n"
                f"• Tổng: {stats['total']} học sinh\n"
                f"• Đã nhận xét: {stats['filled']} ô\n"
                f"• Bỏ qua: {stats['skipped']} ô\n\n"
                f"Nhấn 'XUẤT FILE' để lưu kết quả.")

        except Exception as e:
            self._vnedu_log(f"LỖI: {str(e)}")
            messagebox.showerror("Lỗi", f"Lỗi xử lý:\n{str(e)}")

    def _open_subject_config(self):
        """Mở popup cấu hình nhận xét môn học (rộng rãi, dễ dùng)"""
        if self._subject_config_win and self._subject_config_win.winfo_exists():
            self._subject_config_win.focus()
            return

        win = ctk.CTkToplevel(self)
        self._subject_config_win = win
        win.title("⚙ Cấu Hình Nhận Xét Môn Học — ETA Insight")
        win.geometry("1050x720")
        win.configure(fg_color="#FFF8F0")
        win.transient(self)
        win.grab_set()
        # Auto-save khi đóng popup
        win.protocol("WM_DELETE_WINDOW", self._cfg_on_close)

        # === Sidebar ===
        sidebar = ctk.CTkFrame(win, width=220, fg_color="#FFF0E0", corner_radius=0)
        sidebar.pack(side="left", fill="y")
        sidebar.pack_propagate(False)

        ctk.CTkLabel(sidebar, text="⚙ CẤU HÌNH\nNHẬN XÉT MÔN", font=("Arial", 15, "bold"),
                     text_color="#333").pack(pady=(20,5))
        ctk.CTkLabel(sidebar, text="Chọn cấp học bên dưới\nrồi tùy chỉnh mẫu nhận xét",
                     font=("Arial", 10), text_color="#888").pack(pady=(0,15))

        # Grade buttons
        grade_options = [
            ("🏫 Tiểu học\n(TT27/2020)", "tieu_hoc"),
            ("📖 THCS\n(TT22/2021)", "thcs"),
            ("🎓 THPT\n(TT22/2021)", "thpt"),
        ]
        self._cfg_grade_btns = []
        for label, key in grade_options:
            btn = ctk.CTkButton(sidebar, text=label,
                                fg_color="#E67E22" if key == self._subj_grade_key else "#FFF0E0",
                                text_color="#FFF" if key == self._subj_grade_key else "#333",
                                hover_color="#FFE0B2", height=50, font=("Arial", 12),
                                command=lambda k=key: self._cfg_switch_grade(k))
            btn.pack(fill="x", padx=10, pady=3)
            self._cfg_grade_btns.append((btn, key))

        # Bottom buttons
        bottom = ctk.CTkFrame(sidebar, fg_color="transparent")
        bottom.pack(side="bottom", fill="x", padx=10, pady=15)
        ctk.CTkButton(bottom, text="💾 Lưu", fg_color="#27AE60", hover_color="#2ECC71",
                       width=90, height=32, font=("Arial", 12, "bold"),
                       command=self._cfg_save).pack(side="left", padx=2)
        ctk.CTkButton(bottom, text="↩ Reset", fg_color="#E67E22", hover_color="#F39C12",
                       width=90, height=32, font=("Arial", 11),
                       command=self._cfg_reset).pack(side="left", padx=2)

        # === Main content (scrollable) ===
        self._cfg_content = ctk.CTkScrollableFrame(win, fg_color="#FFFFFF", corner_radius=0)
        self._cfg_content.pack(side="right", fill="both", expand=True)

        # Status label
        self._cfg_status = ctk.CTkLabel(win, text="", font=("Arial", 11), text_color="#27AE60")
        self._cfg_status.place(relx=0.95, y=10, anchor="ne")

        # Build content
        self._cfg_build_content()

    def _cfg_switch_grade(self, grade_key):
        """Chuyển cấp học trong popup"""
        if grade_key == self._subj_grade_key:
            return

        ok = messagebox.askyesno("Đổi cấp học",
            "Chuyển sang preset cấp mới?\n\nCác mẫu nhận xét sẽ được thay thế.\nBấm 'Yes' để đổi.")
        if not ok:
            return

        self._subj_grade_key = grade_key

        # Update button highlight
        for btn, key in self._cfg_grade_btns:
            if key == grade_key:
                btn.configure(fg_color="#E67E22", text_color="#FFF")
            else:
                btn.configure(fg_color="#FFF0E0", text_color="#333")

        preset = get_preset_as_settings(grade_key)
        self._cfg_build_content(preset)
        self._cfg_show_status("Đã chuyển cấp học!")

    def _cfg_build_content(self, settings=None):
        """Build nội dung popup (gọi khi init hoặc đổi cấp)"""
        for w in self._cfg_content.winfo_children():
            w.destroy()
        self.subj_level_widgets = []
        self.subj_text_widgets = {}

        if settings is None:
            settings = load_subject_settings()

        grade_labels = {"tieu_hoc": "TIỂU HỌC (TT27/2020)", "thcs": "THCS (TT22/2021)", "thpt": "THPT (TT22/2021)"}
        grade_lbl = grade_labels.get(self._subj_grade_key, "")

        ctk.CTkLabel(self._cfg_content, text=f"📊 MỨC ĐIỂM SỐ — {grade_lbl}",
                     font=("Arial", 16, "bold"), text_color="#333").pack(pady=(15,5), padx=15, anchor="w")
        ctk.CTkLabel(self._cfg_content, text="Mỗi dòng = 1 câu nhận xét mẫu. Tên mức có thể sửa tùy ý.",
                     font=("Arial", 11), text_color="#888").pack(padx=15, anchor="w")

        level_colors = ["#27AE60", "#2980B9", "#E67E22", "#E74C3C"]
        num_cfg = settings.get("numeric", {})
        sorted_levels = sorted(num_cfg.items(), key=lambda x: x[1].get("min", 0), reverse=True)

        for idx, (level_key, level_data) in enumerate(sorted_levels):
            color = level_colors[idx % len(level_colors)]
            is_last = idx == len(sorted_levels) - 1

            frame = ctk.CTkFrame(self._cfg_content, fg_color="#FFF8F0", corner_radius=8,
                                  border_width=1, border_color=color)
            frame.pack(fill="x", padx=15, pady=6)

            row = ctk.CTkFrame(frame, fg_color="transparent")
            row.pack(fill="x", padx=12, pady=(8,4))

            name_entry = ctk.CTkEntry(row, width=160, height=30, font=("Arial", 13, "bold"),
                                       text_color=color, border_width=1, border_color=color)
            name_entry.insert(0, level_data.get("name", level_key))
            name_entry.pack(side="left")

            if not is_last:
                ctk.CTkLabel(row, text="  ≥  ", font=("Arial", 13, "bold"), text_color=color).pack(side="left")
                min_entry = ctk.CTkEntry(row, width=55, height=30, font=("Arial", 13),
                                          justify="center", border_width=1, border_color=color)
                min_entry.insert(0, str(level_data.get("min", 0)))
                min_entry.pack(side="left")
                ctk.CTkLabel(row, text="điểm", font=("Arial", 11), text_color="#888").pack(side="left", padx=5)
            else:
                ctk.CTkLabel(row, text="  < ngưỡng trên", font=("Arial", 11),
                             text_color="#888").pack(side="left", padx=5)
                min_entry = None

            templates = level_data.get("templates", [])
            ctk.CTkLabel(row, text=f"({len(templates)} mẫu)", font=("Arial", 10),
                         text_color="#AAA").pack(side="right", padx=5)

            txt_box = ctk.CTkTextbox(frame, height=90, font=("Arial", 12), border_width=1,
                                      border_color="#E0E0E0", corner_radius=4)
            txt_box.pack(fill="x", padx=12, pady=(0,8))
            txt_box.insert("1.0", "\n".join(templates))

            self.subj_level_widgets.append({
                "key": level_key, "name_entry": name_entry,
                "min_entry": min_entry, "txt_box": txt_box
            })

        # Separator
        ctk.CTkFrame(self._cfg_content, height=2, fg_color="#E67E22").pack(fill="x", padx=15, pady=12)

        # Text-based assessment
        ctk.CTkLabel(self._cfg_content, text="📝 MỨC ĐÁNH GIÁ BẰNG CHỮ (Đ/CĐ)",
                     font=("Arial", 16, "bold"), text_color="#333").pack(pady=(0,5), padx=15, anchor="w")
        ctk.CTkLabel(self._cfg_content, text="Dùng cho các môn đánh giá bằng nhận xét (GDTC, Nghệ thuật, HĐTN...)",
                     font=("Arial", 11), text_color="#888").pack(padx=15, anchor="w")

        text_cfg = settings.get("text", {})
        for text_key, color_val in [("dat", "#27AE60"), ("chuadat", "#E74C3C")]:
            text_data = text_cfg.get(text_key, {})

            frame = ctk.CTkFrame(self._cfg_content, fg_color="#FFF8F0", corner_radius=8,
                                  border_width=1, border_color=color_val)
            frame.pack(fill="x", padx=15, pady=6)

            trow = ctk.CTkFrame(frame, fg_color="transparent")
            trow.pack(fill="x", padx=12, pady=(8,4))

            t_name_entry = ctk.CTkEntry(trow, width=160, height=30, font=("Arial", 13, "bold"),
                                         text_color=color_val, border_width=1, border_color=color_val)
            t_name_entry.insert(0, text_data.get("name", text_key))
            t_name_entry.pack(side="left")

            t_templates = text_data.get("templates", [])
            ctk.CTkLabel(trow, text=f"({len(t_templates)} mẫu)", font=("Arial", 10),
                         text_color="#AAA").pack(side="right", padx=5)

            t_txt_box = ctk.CTkTextbox(frame, height=70, font=("Arial", 12), border_width=1,
                                        border_color="#E0E0E0", corner_radius=4)
            t_txt_box.pack(fill="x", padx=12, pady=(0,8))
            t_txt_box.insert("1.0", "\n".join(t_templates))

            self.subj_text_widgets[text_key] = {
                "name_entry": t_name_entry, "txt_box": t_txt_box,
                "values": text_data.get("values", [])
            }

    def _cfg_save(self):
        """Lưu settings từ popup"""
        self._subj_collect_settings()
        save_subject_settings(self.subject_proc.settings)
        self._cfg_show_status("💾 Đã lưu cấu hình!")
        self._vnedu_log("💾 Đã lưu cấu hình nhận xét môn học!")

    def _cfg_reset(self):
        """Reset về mặc định trong popup"""
        ok = messagebox.askyesno("Xác nhận", "Reset toàn bộ về mặc định?\nCác thay đổi sẽ bị mất!")
        if not ok:
            return
        preset = get_preset_as_settings(self._subj_grade_key)
        self._cfg_build_content(preset)
        self.subject_proc.settings = preset
        save_subject_settings(preset)
        self._cfg_show_status("↩ Đã reset về mặc định!")
        self._vnedu_log("↩ Đã reset nhận xét về mặc định!")

    def _cfg_show_status(self, text):
        """Hiện status tạm trong popup"""
        try:
            self._cfg_status.configure(text=text)
            self._subject_config_win.after(3000, lambda: self._cfg_status.configure(text=""))
        except Exception:
            pass

    def _cfg_on_close(self):
        """Auto-save settings khi đóng popup"""
        try:
            if self.subj_level_widgets:
                self._subj_collect_settings()
                save_subject_settings(self.subject_proc.settings)
                self._vnedu_log("💾 Tự động lưu cấu hình khi đóng popup!")
        except Exception:
            pass
        # Cleanup widget refs
        self.subj_level_widgets = []
        self.subj_text_widgets = {}
        # Đóng popup
        self._subject_config_win.destroy()

    def _auto_detect_grade(self, info):
        """Auto-detect cấp học từ thông tin file sổ điểm"""
        import re

        school = (info.get("school", "") or "").upper()
        class_name = (info.get("class", "") or "").upper()
        detected = None

        # Ưu tiên 1: Từ tên trường
        if any(kw in school for kw in ["TIỂU HỌC", "TIEU HOC", "TH "]):
            detected = "tieu_hoc"
        elif any(kw in school for kw in ["THPT", "TRUNG HỌC PHỔ THÔNG", "PHỔ THÔNG"]):
            detected = "thpt"
        elif any(kw in school for kw in ["THCS", "TRUNG HỌC CƠ SỞ"]):
            detected = "thcs"

        # Ưu tiên 2: Từ tên lớp (Lớp 1-5 = TH, 6-9 = THCS, 10-12 = THPT)
        if not detected:
            match = re.search(r'(?:LỚP|LOP|KHỐI|KHOI)\s*(\d+)', class_name)
            if match:
                grade_num = int(match.group(1))
                if 1 <= grade_num <= 5:
                    detected = "tieu_hoc"
                elif 6 <= grade_num <= 9:
                    detected = "thcs"
                elif 10 <= grade_num <= 12:
                    detected = "thpt"

        if detected and detected != self._subj_grade_key:
            grade_names = {"tieu_hoc": "Tiểu học (TT27)", "thcs": "THCS (TT22)", "thpt": "THPT (TT22)"}
            self._subj_grade_key = detected
            preset = get_preset_as_settings(detected)
            self.subject_proc.settings = preset
            save_subject_settings(preset)
            self._vnedu_log(f"🎯 Auto-detect: {grade_names.get(detected, detected)}")
        elif not detected:
            self._vnedu_log(f"ℹ Cấp học: dùng cấu hình hiện tại")

    def _subj_collect_settings(self):
        """Thu thập settings từ UI popup (nếu đang mở) hoặc load từ file"""
        # Nếu popup đã đóng hoặc chưa mở → dùng settings từ file
        if not self.subj_level_widgets:
            saved = load_subject_settings()
            if saved and saved.get("numeric"):
                self.subject_proc.settings = saved
            return

        def get_lines(textbox):
            try:
                return [l.strip() for l in textbox.get("1.0", "end").strip().split("\n") if l.strip()]
            except Exception:
                return []

        try:
            numeric = {}
            for w in self.subj_level_widgets:
                name = w["name_entry"].get().strip()
                min_val = float(w["min_entry"].get()) if w["min_entry"] else 0
                templates = get_lines(w["txt_box"])
                numeric[w["key"]] = {"name": name, "min": min_val, "templates": templates}

            text = {}
            for text_key, tw in self.subj_text_widgets.items():
                name = tw["name_entry"].get().strip()
                templates = get_lines(tw["txt_box"])
                text[text_key] = {"name": name, "values": tw.get("values", []), "templates": templates}

            self.subject_proc.settings = {"numeric": numeric, "text": text, "grade": self._subj_grade_key}
        except (ValueError, AttributeError, Exception):
            # Fallback: load từ file
            saved = load_subject_settings()
            if saved and saved.get("numeric"):
                self.subject_proc.settings = saved

    def _vnedu_show_preview(self):
        """Hiển thị preview file VNEDU sử dụng Treeview để tối ưu hiệu năng"""
        # Xóa các widget hiện tại trong frame
        for w in self.vnedu_preview_frame.winfo_children():
            w.destroy()

        data = self.vnedu.get_preview_data(max_rows=100) # Cho phép hiện nhiều hơn vì Treeview rất nhẹ
        if not data:
            return

        headers = data["headers"]
        rows = data["rows"]

        # Chọn cột quan trọng (STT, Tên, Mức đạt được, Điểm)
        key_cols = []
        for j, h in enumerate(headers):
            if j < 4 or "Mức" in h or "Điểm" in h or j >= 25:
                key_cols.append((j, h))
        key_cols = key_cols[:15]  # Tối đa 15 cột để không quá chật

        col_ids = [str(i) for i in range(len(key_cols))]
        
        # Setup Treeview
        self.vnedu_tree = ttk.Treeview(self.vnedu_preview_frame, columns=col_ids, show="headings", style="Treeview")
        
        # Scrollbars
        y_scroll = ttk.Scrollbar(self.vnedu_preview_frame, orient="vertical", command=self.vnedu_tree.yview)
        x_scroll = ttk.Scrollbar(self.vnedu_preview_frame, orient="horizontal", command=self.vnedu_tree.xview)
        self.vnedu_tree.configure(yscrollcommand=y_scroll.set, xscrollcommand=x_scroll.set)

        y_scroll.pack(side="right", fill="y")
        x_scroll.pack(side="bottom", fill="x")
        self.vnedu_tree.pack(fill="both", expand=True)

        # Cấu hình Columns & Headings
        for i, (orig_j, h_text) in enumerate(key_cols):
            self.vnedu_tree.heading(str(i), text=h_text[:20])
            width = 150 if i == 2 else 70 # Cột Tên (index 2) rộng hơn
            self.vnedu_tree.column(str(i), width=width, minwidth=50, anchor="center")

        # Thêm Data
        for row in rows:
            display_row = []
            for (orig_j, _) in key_cols:
                val = row[orig_j] if orig_j < len(row) else ""
                display_row.append(val[:30])
            self.vnedu_tree.insert("", "end", values=display_row)

        # Row count
        valid_rows = sum(1 for r in rows if any(r))
        ctk.CTkLabel(self.vnedu_preview_frame, text=f"Hiển thị {valid_rows} dòng dữ liệu mẫu",
                     font=("Arial", 10), text_color="#7F8C8D").pack(anchor="e", padx=5, pady=(2,0))

    def _vnedu_run(self):
        """Chạy xử lý VNEDU: điền mức đạt được"""
        if not self.vnedu_loaded_file:
            messagebox.showwarning("Chưa có file", "Vui lòng tải file VNEDU trước!")
            return

        # Cập nhật settings từ UI
        try:
            t_min = float(self.vnedu_t_min.get())
            h_min = float(self.vnedu_h_min.get())
            self.vnedu.settings = {"score_T_min": t_min, "score_H_min": h_min}
        except ValueError:
            pass

        self._vnedu_log("Bắt đầu điền mức đạt được...")
        try:
            stats = self.vnedu.process()
            self._vnedu_log(f"✅ Hoàn tất! {stats['total']} học sinh")
            self._vnedu_log(f"   Đã điền: {stats['filled']} ô mức đạt được")
            self._vnedu_log(f"   Bỏ qua (đã có): {stats['skipped']} ô")

            for detail in stats.get("details", [])[:10]:
                self._vnedu_log(f"   {detail}")

            self.vnedu_export_btn.configure(state="normal")

            # Cập nhật preview
            self._vnedu_show_preview()
            messagebox.showinfo("Thành công",
                f"Đã điền mức đạt được tự động!\n\n"
                f"• Tổng: {stats['total']} học sinh\n"
                f"• Đã điền: {stats['filled']} ô\n"
                f"• Bỏ qua: {stats['skipped']} ô\n\n"
                f"Nhấn 'XUẤT FILE' để lưu kết quả.")

        except Exception as e:
            self._vnedu_log(f"LỖI: {str(e)}")
            messagebox.showerror("Lỗi", f"Lỗi xử lý:\n{str(e)}")

    def _vnedu_export(self):
        """Xuất file VNEDU đã xử lý (cả 2 mode)"""
        if self._vnedu_mode == "subject":
            if not self.subject_proc.wb:
                return
        else:
            if not self.vnedu.wb:
                return

        default_name = os.path.splitext(os.path.basename(self.vnedu_loaded_file))[0] + "_DA_XU_LY.xlsx"
        output_path = filedialog.asksaveasfilename(
            title="Lưu file kết quả",
            defaultextension=".xlsx",
            initialfile=default_name,
            filetypes=[("Excel files", "*.xlsx")]
        )
        if output_path:
            try:
                if self._vnedu_mode == "subject":
                    self.subject_proc.save_output(output_path)
                else:
                    self.vnedu.save_output(output_path)
                self._vnedu_log(f"💾 Đã xuất: {os.path.basename(output_path)}")
                messagebox.showinfo("Thành công", f"Đã xuất file thành công!\n{output_path}")
            except Exception as e:
                self._vnedu_log(f"LỖI xuất: {str(e)}")
                messagebox.showerror("Lỗi", f"Không thể lưu:\n{str(e)}")

    # =======================================
    # TRANG CHUYỂN ĐỔI (CONVERT)
    # =======================================
    def _build_page_convert(self, parent):
        """Xây dựng trang Chuyển đổi VNEDU ↔ CSDL Ngành"""
        paned = tk.PanedWindow(parent, orient="horizontal", sashwidth=6,
                               bg="#D5C9B8", sashrelief="flat")
        paned.pack(fill="both", expand=True, padx=10, pady=(10,5))

        # === LEFT: Upload & Actions ===
        left_wrapper = ctk.CTkFrame(paned, fg_color=BG_CARD, corner_radius=12,
                                     border_width=1, border_color="#E0D5C5")
        paned.add(left_wrapper, minsize=400, stretch="always")
        left = ctk.CTkScrollableFrame(left_wrapper, fg_color=BG_CARD, corner_radius=0)
        left.pack(fill="both", expand=True)

        # Title
        title_frame = ctk.CTkFrame(left, fg_color="#8E44AD", corner_radius=8)
        title_frame.pack(fill="x", padx=20, pady=(20,10))
        ctk.CTkLabel(title_frame, text="🔄 CHUYỂN ĐỔI DỮ LIỆU",
                     font=("Arial", 15, "bold"), text_color="white").pack(padx=15, pady=8, anchor="w")
        ctk.CTkLabel(title_frame, text="Copy điểm & đánh giá giữa VNEDU ↔ CSDL Ngành",
                     font=("Arial", 11), text_color="#D2B4DE").pack(padx=15, pady=(0,8), anchor="w")

        # Section 1: File Nguồn
        s1 = ctk.CTkFrame(left, fg_color="transparent")
        s1.pack(fill="x", padx=20, pady=(5,5))
        ctk.CTkLabel(s1, text="1. FILE NGUỒN (đã có dữ liệu)", font=("Arial", 13, "bold"),
                     text_color=TEXT_DARK).pack(anchor="w")
        ctk.CTkLabel(s1, text="File VNEDU hoặc CSDL Ngành đã chấm điểm.",
                     font=("Arial", 11), text_color=TEXT_MID).pack(anchor="w", pady=(2,8))

        ctk.CTkButton(s1, text="📂 Chọn File Nguồn...", fg_color="#FFFFFF",
                       text_color="#8E44AD", border_width=1, border_color="#8E44AD",
                       hover_color="#F4ECF7", font=("Arial", 12, "bold"),
                       height=36, width=200, command=self._convert_open_source).pack(anchor="w")

        self.cvt_source_label = ctk.CTkLabel(s1, text="Chưa chọn file", font=("Arial", 11, "italic"),
                                              text_color=TEXT_MID)
        self.cvt_source_label.pack(anchor="w", pady=(5,0))
        self.cvt_source_info = ctk.CTkLabel(s1, text="", font=("Arial", 10), text_color="#7D3C98")
        self.cvt_source_info.pack(anchor="w")

        ctk.CTkFrame(left, height=1, fg_color="#EAECEE").pack(fill="x", padx=20, pady=8)

        # Section 2: File Đích
        s2 = ctk.CTkFrame(left, fg_color="transparent")
        s2.pack(fill="x", padx=20, pady=5)
        ctk.CTkLabel(s2, text="2. FILE ĐÍCH (cần điền dữ liệu vào)", font=("Arial", 13, "bold"),
                     text_color=TEXT_DARK).pack(anchor="w")
        ctk.CTkLabel(s2, text="File trống hoặc chưa hoàn chỉnh của hệ thống kia.",
                     font=("Arial", 11), text_color=TEXT_MID).pack(anchor="w", pady=(2,8))

        ctk.CTkButton(s2, text="📂 Chọn File Đích...", fg_color="#FFFFFF",
                       text_color="#8E44AD", border_width=1, border_color="#8E44AD",
                       hover_color="#F4ECF7", font=("Arial", 12, "bold"),
                       height=36, width=200, command=self._convert_open_dest).pack(anchor="w")

        self.cvt_dest_label = ctk.CTkLabel(s2, text="Chưa chọn file", font=("Arial", 11, "italic"),
                                            text_color=TEXT_MID)
        self.cvt_dest_label.pack(anchor="w", pady=(5,0))
        self.cvt_dest_info = ctk.CTkLabel(s2, text="", font=("Arial", 10), text_color="#7D3C98")
        self.cvt_dest_info.pack(anchor="w")

        ctk.CTkFrame(left, height=1, fg_color="#EAECEE").pack(fill="x", padx=20, pady=8)

        # Section 3: Hướng chuyển đổi (auto-detect)
        s3 = ctk.CTkFrame(left, fg_color="transparent")
        s3.pack(fill="x", padx=20, pady=5)
        ctk.CTkLabel(s3, text="3. HƯỚNG CHUYỂN ĐỔI", font=("Arial", 13, "bold"),
                     text_color=TEXT_DARK).pack(anchor="w")

        self.cvt_direction_frame = ctk.CTkFrame(s3, fg_color="#FFFFFF", border_width=1,
                                                  border_color="#E0E0E0", corner_radius=6)
        self.cvt_direction_frame.pack(fill="x", pady=5)
        self.cvt_direction_label = ctk.CTkLabel(self.cvt_direction_frame,
                                                 text="Chọn 2 file để xác định hướng chuyển đổi...",
                                                 font=("Arial", 12), text_color=TEXT_MID)
        self.cvt_direction_label.pack(padx=12, pady=10)

        ctk.CTkFrame(left, height=1, fg_color="#EAECEE").pack(fill="x", padx=20, pady=5)

        # Section 4: Actions
        s4 = ctk.CTkFrame(left, fg_color="transparent")
        s4.pack(fill="x", padx=20, pady=(5,15))
        ctk.CTkLabel(s4, text="4. THỰC THI", font=("Arial", 13, "bold"),
                     text_color=TEXT_DARK).pack(anchor="w", pady=(0,8))

        self.cvt_run_btn = ctk.CTkButton(s4, text="🔄 CHUYỂN ĐỔI DỮ LIỆU",
                                          fg_color="#8E44AD", hover_color="#7D3C98",
                                          font=("Arial", 12, "bold"), height=40,
                                          command=self._convert_run, state="disabled")
        self.cvt_run_btn.pack(fill="x", pady=(0,8))

        self.cvt_export_btn = ctk.CTkButton(s4, text="💾 XUẤT FILE KẾT QUẢ",
                                             fg_color=SUCCESS, hover_color="#219A52",
                                             font=("Arial", 12, "bold"), height=40,
                                             command=self._convert_export, state="disabled")
        self.cvt_export_btn.pack(fill="x")

        # === RIGHT: Preview & Log ===
        right = ctk.CTkFrame(paned, fg_color=BG_CARD, corner_radius=12,
                             border_width=1, border_color="#E0D5C5")
        paned.add(right, minsize=350, stretch="always")

        # Preview header
        preview_header = ctk.CTkFrame(right, fg_color="#6C3483", corner_radius=8, height=40)
        preview_header.pack(fill="x", padx=12, pady=(12,0))
        preview_header.pack_propagate(False)
        ctk.CTkLabel(preview_header, text="📋 KẾT QUẢ CHUYỂN ĐỔI",
                     font=("Arial", 13, "bold"), text_color="white").pack(side="left", padx=15)
        self.cvt_preview_stats = ctk.CTkLabel(preview_header, text="",
                                               font=("Arial", 10), text_color="#D7BDE2")
        self.cvt_preview_stats.pack(side="right", padx=15)

        # Preview
        self.cvt_preview_frame = ctk.CTkFrame(right, fg_color="#FFFFFF", corner_radius=0)
        self.cvt_preview_frame.pack(fill="both", expand=True, padx=12, pady=(5,5))

        self.cvt_ph = ctk.CTkFrame(self.cvt_preview_frame, fg_color="transparent")
        self.cvt_ph.pack(fill="both", expand=True, pady=40)
        ctk.CTkLabel(self.cvt_ph, text="🔄", font=("Arial", 42), text_color="#BDC3C7").pack()
        ctk.CTkLabel(self.cvt_ph, text="Chọn 2 file và bấm Chuyển Đổi để xem kết quả",
                     font=("Arial", 13), text_color="#7F8C8D").pack(pady=(10,0))

        self.cvt_tree = None

        # Log
        log_header = ctk.CTkFrame(right, fg_color="#6C3483", corner_radius=8, height=32)
        log_header.pack(fill="x", padx=12, pady=(5,0))
        log_header.pack_propagate(False)
        ctk.CTkLabel(log_header, text="📝 NHẬT KÝ CHUYỂN ĐỔI",
                     font=("Arial", 11, "bold"), text_color="white").pack(side="left", padx=15)

        self.cvt_log_box = ctk.CTkTextbox(right, height=150, fg_color="#1A252F",
                                           text_color="#BB8FCE", font=("Consolas", 11),
                                           corner_radius=0, border_width=1, border_color="#34495E")
        self.cvt_log_box.pack(fill="x", padx=12, pady=(0,12))
        self._convert_log("Chọn File Nguồn và File Đích để bắt đầu!")

    def _convert_log(self, text):
        self.cvt_log_box.insert("end", f"→ {text}\n")
        self.cvt_log_box.see("end")

    def _convert_open_source(self):
        """Chọn file nguồn"""
        filepath = filedialog.askopenfilename(
            title="Chọn File Nguồn (đã có dữ liệu)",
            filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")]
        )
        if not filepath:
            return
        try:
            self.converter = ConverterProcessor()  # Reset
            info = self.converter.load_source(filepath)
            fname = os.path.basename(filepath)
            self.cvt_source_label.configure(text=f"✅ {fname}", text_color=SUCCESS)
            self.cvt_source_info.configure(text=f"Loại: {info['type']} | {info['total_students']} học sinh")
            self._convert_log(f"Nguồn: {fname} ({info['type']}, {info['total_students']} HS)")
            self._update_convert_direction()
        except Exception as e:
            self._convert_log(f"LỖI: {str(e)}")
            messagebox.showerror("Lỗi", str(e))

    def _convert_open_dest(self):
        """Chọn file đích"""
        filepath = filedialog.askopenfilename(
            title="Chọn File Đích (cần điền dữ liệu vào)",
            filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")]
        )
        if not filepath:
            return
        try:
            info = self.converter.load_dest(filepath)
            fname = os.path.basename(filepath)
            self.cvt_dest_label.configure(text=f"✅ {fname}", text_color=SUCCESS)
            self.cvt_dest_info.configure(text=f"Loại: {info['type']} | {info['total_students']} học sinh")
            self._convert_log(f"Đích: {fname} ({info['type']}, {info['total_students']} HS)")
            self._update_convert_direction()
        except Exception as e:
            self._convert_log(f"LỖI: {str(e)}")
            messagebox.showerror("Lỗi", str(e))

    def _update_convert_direction(self):
        """Cập nhật hiển thị hướng chuyển đổi"""
        src = self.converter.source_type
        dst = self.converter.dest_type
        if src and dst:
            direction = f"{src.upper()}  →  {dst.upper()}"
            self.cvt_direction_label.configure(
                text=f"📌 {direction}",
                text_color="#8E44AD", font=("Arial", 14, "bold")
            )
            self.cvt_run_btn.configure(state="normal")
        elif src:
            expected = "CSDL Ngành" if src == "vnedu" else "VNEDU"
            self.cvt_direction_label.configure(
                text=f"Nguồn: {src.upper()} → Chờ chọn file đích ({expected})...",
                text_color=TEXT_MID
            )

    def _convert_run(self):
        """Thực hiện chuyển đổi"""
        if not self.converter.source_wb or not self.converter.dest_wb:
            messagebox.showwarning("Thiếu file", "Vui lòng chọn cả File Nguồn và File Đích!")
            return

        self._convert_log("Bắt đầu chuyển đổi...")
        try:
            stats = self.converter.convert()
            # In logs
            for line in stats.get("details", []):
                self._convert_log(line)

            self.cvt_preview_stats.configure(
                text=f"✅ {stats['matched']}/{stats['total_source']} HS | {stats['cells_filled']} ô"
            )
            self.cvt_export_btn.configure(state="normal")

            # Hiển thị preview file đích đã điền
            self._convert_show_preview()

            messagebox.showinfo("Hoàn tất",
                f"Chuyển đổi thành công!\n\n"
                f"Hướng: {stats['direction']}\n"
                f"Đã ghép: {stats['matched']}/{stats['total_source']} học sinh\n"
                f"Ô đã điền: {stats['cells_filled']}\n"
                f"Không tìm thấy: {stats['not_found']} HS\n\n"
                f"Nhấn 'XUẤT FILE' để lưu kết quả.")
        except Exception as e:
            self._convert_log(f"LỖI: {str(e)}")
            messagebox.showerror("Lỗi", f"Lỗi chuyển đổi:\n{str(e)}")

    def _convert_show_preview(self):
        """Hiển thị preview file đích sau chuyển đổi"""
        for w in self.cvt_preview_frame.winfo_children():
            w.destroy()

        data = self.converter.get_preview_data("dest", max_rows=50)
        if not data:
            return

        headers = data["headers"]
        rows = data["rows"]

        # Chọn cột quan trọng
        key_cols = []
        for j, h in enumerate(headers):
            h_str = str(h)
            if j < 5 or "Mức" in h_str or "Điểm" in h_str:
                key_cols.append((j, h_str))
        key_cols = key_cols[:15]

        col_ids = [str(i) for i in range(len(key_cols))]
        self.cvt_tree = ttk.Treeview(self.cvt_preview_frame, columns=col_ids, show="headings", style="Treeview")

        y_scroll = ttk.Scrollbar(self.cvt_preview_frame, orient="vertical", command=self.cvt_tree.yview)
        x_scroll = ttk.Scrollbar(self.cvt_preview_frame, orient="horizontal", command=self.cvt_tree.xview)
        self.cvt_tree.configure(yscrollcommand=y_scroll.set, xscrollcommand=x_scroll.set)
        y_scroll.pack(side="right", fill="y")
        x_scroll.pack(side="bottom", fill="x")
        self.cvt_tree.pack(fill="both", expand=True)

        for i, (orig_j, h_text) in enumerate(key_cols):
            self.cvt_tree.heading(str(i), text=h_text[:20])
            width = 150 if i in (2, 3) else 70
            self.cvt_tree.column(str(i), width=width, minwidth=50, anchor="center")

        for row in rows:
            if not any(row):
                continue
            display_row = []
            for (orig_j, _) in key_cols:
                val = row[orig_j] if orig_j < len(row) else ""
                display_row.append(val[:30])
            self.cvt_tree.insert("", "end", values=display_row)

    def _convert_export(self):
        """Xuất file đã chuyển đổi"""
        if not self.converter.dest_wb:
            return
        src_name = os.path.splitext(self.converter.source_path or "output")[0]
        dest_type = (self.converter.dest_type or "").upper()
        default_name = f"{os.path.basename(src_name)}_SANG_{dest_type}.xlsx"
        output_path = filedialog.asksaveasfilename(
            title="Lưu file đã chuyển đổi",
            defaultextension=".xlsx",
            initialfile=default_name,
            filetypes=[("Excel files", "*.xlsx")]
        )
        if output_path:
            try:
                self.converter.save_output(output_path)
                self._convert_log(f"💾 Đã xuất: {os.path.basename(output_path)}")
                messagebox.showinfo("Thành công", f"Đã xuất file thành công!\n{output_path}")
            except Exception as e:
                self._convert_log(f"LỖI xuất: {str(e)}")
                messagebox.showerror("Lỗi", f"Không thể lưu:\n{str(e)}")

    # === COLOR MAP cho mức đánh giá ===
    LEVEL_COLORS = {
        "T": ("#27AE60", "#E8F8F5"),   # Tốt → xanh lá
        "T ": ("#27AE60", "#E8F8F5"),
        "H": ("#E67E22", "#FEF5E7"),   # Hoàn thành → cam
        "H ": ("#E67E22", "#FEF5E7"),
        "Đ": ("#E67E22", "#FEF5E7"),   # Đạt → cam
        "Đ ": ("#E67E22", "#FEF5E7"),
        "C": ("#E74C3C", "#FDEDEC"),   # Chưa HT → đỏ
        "C ": ("#E74C3C", "#FDEDEC"),
        "XS": ("#8E44AD", "#F5EEF8"),  # Xuất sắc → tím
        "K": ("#2980B9", "#EBF5FB"),   # Khá → xanh dương
        "G": ("#27AE60", "#E8F8F5"),   # Giỏi → xanh
        "TB": ("#F39C12", "#FEF9E7"),  # Trung bình → vàng
        "Y": ("#E74C3C", "#FDEDEC"),   # Yếu → đỏ
    }

    def _log(self, text):
        self.log_box.insert("end", f"→ {text}\n")
        self.log_box.see("end")

    def _open_file(self):
        filepath = filedialog.askopenfilename(
            title="Chọn file Excel đánh giá học sinh",
            filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")]
        )
        if not filepath:
            return

        self.loaded_file = filepath
        filename = os.path.basename(filepath)
        self.file_label.configure(text=f"✅ {filename}", text_color=SUCCESS)
        self._log(f"Đã tải file: {filename}")

        try:
            file_type = self.processor.load_file(filepath)
            type_display = {"nlpc": "Năng lực Phẩm chất (Tiểu học)",
                           "dinhky_monhoc": "Đánh giá định kỳ theo môn",
                           "unknown": "Chưa xác định"}.get(file_type, file_type)
            self._log(f"Loại file: {type_display}")

            sheets = self.processor.get_sheet_info()
            info_text = f"Loại file: {type_display}\n"
            total_students = 0
            for s in sheets:
                info_text += f"  • Sheet '{s['name']}': {s['rows']} học sinh\n"
                total_students += s['rows']
            self.info_label.configure(text=info_text, text_color=TEXT_DARK)

            if file_type == "nlpc":
                self.cap_display_var.set("Tiểu Học")
                self._log("Tự động chọn: Tiểu Học (file NLPC)")

            # Lưu sheets để chuyển tab
            self._current_sheets = sheets
            self._current_sheet_idx = 0
            self.preview_stats.configure(text=f"📊 {len(sheets)} sheet • {total_students} học sinh")
            self._build_sheet_tabs(sheets)
            self._show_preview_sheet(sheets[0]["name"])

            self.run_btn.configure(state="normal")
            self._log("Sẵn sàng điền nhận xét!")

        except Exception as e:
            self._log(f"LỖI: {str(e)}")
            messagebox.showerror("Lỗi", f"Không thể đọc file:\n{str(e)}")

    def _build_sheet_tabs(self, sheets):
        """Tạo các tab chọn sheet"""
        for w in self.sheet_tabs_frame.winfo_children():
            w.destroy()

        self._tab_buttons = []
        for i, s in enumerate(sheets):
            btn = ctk.CTkButton(
                self.sheet_tabs_frame, text=f"  {s['name']}  ",
                font=("Arial", 11), height=28, corner_radius=5,
                fg_color=ACCENT if i == 0 else "transparent",
                text_color="white" if i == 0 else TEXT_DARK,
                hover_color=ACCENT_HOVER,
                command=lambda idx=i, name=s['name']: self._switch_sheet_tab(idx, name)
            )
            btn.pack(side="left", padx=(8 if i == 0 else 2, 2), pady=3)
            self._tab_buttons.append(btn)

    def _switch_sheet_tab(self, idx, name):
        """Chuyển tab sheet"""
        self._current_sheet_idx = idx
        for i, btn in enumerate(self._tab_buttons):
            if i == idx:
                btn.configure(fg_color=ACCENT, text_color="white")
            else:
                btn.configure(fg_color="transparent", text_color=TEXT_DARK)
        self._show_preview_sheet(name)

    def _show_preview(self, sheets):
        """Compatibility wrapper"""
        if sheets:
            self._current_sheets = sheets
            self._build_sheet_tabs(sheets)
            self._show_preview_sheet(sheets[0]["name"])

    def _show_preview_sheet(self, sheet_name):
        """Render bảng preview cho 1 sheet sử dụng Treeview để tối ưu hiệu năng"""
        for w in self.preview_frame.winfo_children():
            w.destroy()

        headers, rows = self.processor.get_preview_data(sheet_name, max_rows=100)
        if not headers:
            return

        # Lọc bỏ cột trống
        visible_cols = []
        for j, h in enumerate(headers):
            if h.strip():
                visible_cols.append((j, h))

        # Giới hạn hiển thị tối đa 15 cột
        visible_cols = visible_cols[:15]
        col_ids = [str(i) for i in range(len(visible_cols))]

        # Setup Treeview
        self.preview_tree = ttk.Treeview(self.preview_frame, columns=col_ids, show="headings", style="Treeview")
        
        # Scrollbars
        y_scroll = ttk.Scrollbar(self.preview_frame, orient="vertical", command=self.preview_tree.yview)
        x_scroll = ttk.Scrollbar(self.preview_frame, orient="horizontal", command=self.preview_tree.xview)
        self.preview_tree.configure(yscrollcommand=y_scroll.set, xscrollcommand=x_scroll.set)

        y_scroll.pack(side="right", fill="y")
        x_scroll.pack(side="bottom", fill="x")
        self.preview_tree.pack(fill="both", expand=True)

        # Cấu hình Columns & Headings
        for i, (orig_j, h_text) in enumerate(visible_cols):
            display_h = h_text[:20] + "…" if len(h_text) > 20 else h_text
            self.preview_tree.heading(str(i), text=display_h)
            # Cột chứa nhận xét hoặc nội dung → rộng hơn
            h_lower = h_text.lower()
            if "nội dung" in h_lower or "nhận xét" in h_lower:
                width = 250
                anchor = "w"
            elif i == 1 or "họ" in h_lower or "tên" in h_lower:
                width = 150
                anchor = "w"
            else:
                width = 80
                anchor = "center"
            self.preview_tree.column(str(i), width=width, minwidth=50, anchor=anchor)

        # Thêm Data
        for row in rows:
            if not any(row):
                continue
            display_row = []
            for (orig_j, _) in visible_cols:
                val = row[orig_j] if orig_j < len(row) else ""
                display_row.append(val[:60])
            self.preview_tree.insert("", "end", values=display_row)

        # Row count
        valid_rows = sum(1 for r in rows if any(r))
        ctk.CTkLabel(self.preview_frame, text=f"Hiển thị {valid_rows} dòng dữ liệu mẫu",
                     font=("Arial", 10), text_color="#7F8C8D").pack(anchor="e", padx=5, pady=(2,0))

    def _run_process(self):
        if not self.loaded_file:
            messagebox.showwarning("Chưa có file", "Vui lòng tải file Excel trước!")
            return

        self._log("Bắt đầu điền nhận xét tự động...")
        cap = CAP_REVERSE.get(self.cap_display_var.get(), "tieu_hoc")
        file_type = self.processor.file_type

        try:
            if file_type == "nlpc":
                count = self.processor.process_nlpc(self.cb, cap)
                self._log(f"✅ Đã xử lý NLPC: {count} học sinh")
            elif file_type == "dinhky_monhoc":
                count = self.processor.process_monhoc(self.cb, cap)
                self._log(f"✅ Đã điền nhận xét: {count} ô")
            else:
                # Try both
                count1 = self.processor.process_nlpc(self.cb, cap)
                count2 = self.processor.process_monhoc(self.cb, cap)
                self._log(f"✅ Đã xử lý: {count1} NLPC + {count2} môn học")

            self.export_btn.configure(state="normal")
            self._log("Hoàn tất! Nhấn 'XUẤT FILE KẾT QUẢ' để lưu.")

            # Reload preview để GV xem kết quả ngay
            try:
                sheets = self.processor.get_sheet_info()
                if sheets:
                    self._show_preview(sheets)
                    self.preview_stats.configure(text="✅ Đã điền nhận xét!")
            except Exception:
                pass

            messagebox.showinfo("Thành công", f"Đã điền nhận xét tự động thành công!\nNhấn 'Xuất file' để lưu kết quả.")
        except Exception as e:
            self._log(f"LỖI: {str(e)}")
            messagebox.showerror("Lỗi", f"Lỗi xử lý:\n{str(e)}")

    def _export_file(self):
        if not self.processor.wb:
            return

        default_name = os.path.splitext(os.path.basename(self.loaded_file))[0] + "_DA_NHAN_XET.xlsx"
        output_path = filedialog.asksaveasfilename(
            title="Lưu file kết quả",
            defaultextension=".xlsx",
            initialfile=default_name,
            filetypes=[("Excel files", "*.xlsx")]
        )
        if output_path:
            try:
                self.processor.save_output(output_path)
                self._log(f"💾 Đã xuất file: {os.path.basename(output_path)}")
                messagebox.showinfo("Thành công", f"Đã xuất file thành công!\n{output_path}")
            except Exception as e:
                self._log(f"LỖI xuất file: {str(e)}")
                messagebox.showerror("Lỗi", f"Không thể lưu file:\n{str(e)}")

    def _open_config(self):
        if self.config_win is None or not self.config_win.winfo_exists():
            self.config_win = ConfigWindow(self, self.cb)
        else:
            self.config_win.focus()

    # =======================================
    # AUTO-UPDATE METHODS
    # =======================================
    def _on_update_check_done(self, has_update, info):
        """Callback từ background thread khi kiểm tra xong"""
        try:
            # Log kết quả check update vào nhật ký
            def _log_result():
                for line in get_update_log():
                    self._log(f"🔄 {line}")
                if has_update and info:
                    self._show_update_available(info)
                else:
                    self._log("✅ Đang dùng phiên bản mới nhất!")
            self.after(100, _log_result)
        except Exception:
            pass

    def _show_update_available(self, info):
        """Hiển thị nút cập nhật nhấp nháy trên topbar"""
        self._update_info = info
        self.update_btn.configure(text=f"🔔 Có phiên bản mới {info['version']}!")
        self.update_btn.pack(side="right", padx=(0, 10))
        self._log(f"🔔 Phát hiện phiên bản mới: {info['version']} ({info['file_size_mb']} MB)")
        self._blink_update_btn()

    def _blink_update_btn(self):
        """Hiệu ứng nhấp nháy cho nút cập nhật"""
        if self._update_info is None:
            return
        self._blink_state = not self._blink_state
        if self._blink_state:
            self.update_btn.configure(fg_color="#E74C3C", text_color="white")
        else:
            self.update_btn.configure(fg_color="#F39C12", text_color="white")
        self.after(800, self._blink_update_btn)

    def _on_update_click(self):
        """Khi người dùng nhấn nút cập nhật"""
        info = self._update_info
        if not info:
            return

        notes = info.get("release_notes", "Không có ghi chú.")
        # Giới hạn ghi chú 300 ký tự
        if len(notes) > 300:
            notes = notes[:300] + "..."

        result = messagebox.askyesno(
            f"Cập nhật phiên bản {info['version']}",
            f"Đã có phiên bản mới: {info['version']}\n"
            f"Dung lượng: {info['file_size_mb']} MB\n\n"
            f"📝 Nội dung cập nhật:\n{notes}\n\n"
            f"Bạn có muốn tải và cập nhật ngay không?\n"
            f"(Ứng dụng sẽ tự động khởi động lại sau khi cập nhật)"
        )
        if result:
            self._start_download()

    def _start_download(self):
        """Bắt đầu tải file cập nhật"""
        info = self._update_info
        self._update_info = None  # Dừng nhấp nháy

        # Đổi nút thành progress
        self.update_btn.configure(
            text="⏳ Đang tải 0%...",
            fg_color="#3498DB",
            state="disabled"
        )
        self._log("⏳ Đang tải bản cập nhật...")

        download_update_async(
            info["download_url"],
            progress_callback=self._on_download_progress,
            done_callback=self._on_download_done
        )

    def _on_download_progress(self, downloaded, total):
        """Cập nhật tiến trình tải"""
        percent = int(downloaded / total * 100) if total > 0 else 0
        self.after(0, lambda p=percent: self.update_btn.configure(
            text=f"⏳ Đang tải {p}%..."
        ))

    def _on_download_done(self, file_path):
        """Callback khi tải xong"""
        if file_path and os.path.exists(file_path):
            self.after(0, lambda: self._apply_downloaded_update(file_path))
        else:
            self.after(0, lambda: self._download_failed())

    def _apply_downloaded_update(self, file_path):
        """Áp dụng bản cập nhật đã tải"""
        self.update_btn.configure(text="✅ Tải xong! Đang cập nhật...")
        self._log("✅ Tải xong! Đang áp dụng bản cập nhật...")

        result = messagebox.showinfo(
            "Cập nhật thành công",
            "Đã tải xong bản cập nhật!\n\n"
            "Ứng dụng sẽ tự động khởi động lại.\n"
            "Nhấn OK để tiếp tục."
        )
        apply_update(file_path)

    def _download_failed(self):
        """Xử lý khi tải thất bại"""
        self.update_btn.configure(
            text="❌ Tải thất bại — Thử lại",
            fg_color="#E74C3C",
            state="normal"
        )
        self._log("❌ Tải bản cập nhật thất bại. Kiểm tra kết nối mạng!")
        self._update_info = self._update_info  # Cho phép thử lại


if __name__ == "__main__":
    app = MainApp()
    app.mainloop()
