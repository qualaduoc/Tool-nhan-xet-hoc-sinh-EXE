# manual_column_ui.py - Popup tùy chỉnh cột/dòng cho file Excel bất kỳ
import customtkinter as ctk
from tkinter import messagebox
from openpyxl.utils import column_index_from_string, get_column_letter


def col_letter_to_number(col_str):
    """Chuyển ký hiệu cột Excel (A, S, AB...) sang số (1, 19, 28...)"""
    col_str = col_str.strip().upper()
    if col_str.isdigit():
        return int(col_str)
    try:
        return column_index_from_string(col_str)
    except Exception:
        return None


class ManualColumnConfig:
    """Lưu trữ cấu hình cột/dòng do GV nhập thủ công"""
    def __init__(self):
        self.enabled = False
        self.comment_col = None      # Cột nhận xét (số nguyên, vd: 19 = S)
        self.comment_col_letter = ""  # Hiển thị (vd: "S")
        self.row_start = None        # Dòng bắt đầu (vd: 8)
        self.row_end = None          # Dòng kết thúc (vd: 22)
        self.name_col = None         # (Tùy chọn) Cột họ tên (vd: 2 = B)
        self.name_col_letter = ""
        self.grade_col = None        # (Tùy chọn) Cột mức/điểm tham chiếu
        self.grade_col_letter = ""

    def reset(self):
        self.__init__()


class ManualColumnPopup(ctk.CTkToplevel):
    """Popup cho phép GV nhập cột/dòng thủ công"""

    def __init__(self, parent, config: ManualColumnConfig, on_apply=None, ws=None):
        super().__init__(parent)
        self.config = config
        self.on_apply = on_apply
        self.ws = ws  # worksheet để preview

        self.title("📐 Tùy Chỉnh Vị Trí Cột / Dòng")
        self.geometry("520x540")
        self.configure(fg_color="#FAFAFA")
        self.transient(parent)
        self.grab_set()
        self.resizable(False, False)

        self._build_ui()

    def _build_ui(self):
        # Header
        header = ctk.CTkFrame(self, fg_color="#2C3E50", corner_radius=0, height=55)
        header.pack(fill="x")
        header.pack_propagate(False)
        ctk.CTkLabel(header, text="📐 TÙY CHỈNH VỊ TRÍ CỘT / DÒNG",
                     font=("Arial", 15, "bold"), text_color="white").pack(side="left", padx=20)
        ctk.CTkLabel(header, text="Chính xác 100%",
                     font=("Arial", 11), text_color="#82E0AA").pack(side="right", padx=20)

        # Body
        body = ctk.CTkScrollableFrame(self, fg_color="#FAFAFA")
        body.pack(fill="both", expand=True, padx=15, pady=10)

        # Info text
        ctk.CTkLabel(body, text="Nhập vị trí cột và dòng trong file Excel.",
                     font=("Arial", 12), text_color="#7F8C8D",
                     wraplength=450, justify="left").pack(anchor="w", pady=(0,10))

        # === Section 1: Bắt buộc ===
        sec1 = ctk.CTkFrame(body, fg_color="#FFFFFF", corner_radius=8,
                            border_width=1, border_color="#E0E0E0")
        sec1.pack(fill="x", pady=(0,10))

        ctk.CTkLabel(sec1, text="⚡ BẮT BUỘC", font=("Arial", 13, "bold"),
                     text_color="#E67E22").pack(anchor="w", padx=15, pady=(12,8))

        # Cột nhận xét
        r1 = ctk.CTkFrame(sec1, fg_color="transparent")
        r1.pack(fill="x", padx=15, pady=4)
        ctk.CTkLabel(r1, text="Cột nhận xét:", font=("Arial", 12, "bold"),
                     text_color="#333", width=130, anchor="w").pack(side="left")
        self.col_entry = ctk.CTkEntry(r1, width=80, height=34, font=("Arial", 13),
                                       placeholder_text="VD: S", justify="center")
        self.col_entry.pack(side="left", padx=5)
        ctk.CTkLabel(r1, text="(Nhập chữ cái cột hoặc số)",
                     font=("Arial", 11), text_color="#999").pack(side="left", padx=5)

        # Dòng bắt đầu
        r2 = ctk.CTkFrame(sec1, fg_color="transparent")
        r2.pack(fill="x", padx=15, pady=4)
        ctk.CTkLabel(r2, text="Dòng bắt đầu:", font=("Arial", 12, "bold"),
                     text_color="#333", width=130, anchor="w").pack(side="left")
        self.start_entry = ctk.CTkEntry(r2, width=80, height=34, font=("Arial", 13),
                                         placeholder_text="VD: 8", justify="center")
        self.start_entry.pack(side="left", padx=5)
        ctk.CTkLabel(r2, text="(Dòng học sinh đầu tiên)",
                     font=("Arial", 11), text_color="#999").pack(side="left", padx=5)

        # Dòng kết thúc
        r3 = ctk.CTkFrame(sec1, fg_color="transparent")
        r3.pack(fill="x", padx=15, pady=(4,12))
        ctk.CTkLabel(r3, text="Dòng kết thúc:", font=("Arial", 12, "bold"),
                     text_color="#333", width=130, anchor="w").pack(side="left")
        self.end_entry = ctk.CTkEntry(r3, width=80, height=34, font=("Arial", 13),
                                       placeholder_text="VD: 22", justify="center")
        self.end_entry.pack(side="left", padx=5)
        ctk.CTkLabel(r3, text="(Dòng học sinh cuối cùng)",
                     font=("Arial", 11), text_color="#999").pack(side="left", padx=5)

        # === Section 2: Tùy chọn ===
        sec2 = ctk.CTkFrame(body, fg_color="#FFFFFF", corner_radius=8,
                            border_width=1, border_color="#E0E0E0")
        sec2.pack(fill="x", pady=(0,10))

        ctk.CTkLabel(sec2, text="📎 TÙY CHỌN (không bắt buộc)", font=("Arial", 13, "bold"),
                     text_color="#3498DB").pack(anchor="w", padx=15, pady=(12,8))

        # Cột họ tên
        r4 = ctk.CTkFrame(sec2, fg_color="transparent")
        r4.pack(fill="x", padx=15, pady=4)
        ctk.CTkLabel(r4, text="Cột Họ tên:", font=("Arial", 12),
                     text_color="#333", width=130, anchor="w").pack(side="left")
        self.name_entry = ctk.CTkEntry(r4, width=80, height=34, font=("Arial", 13),
                                        placeholder_text="VD: B", justify="center")
        self.name_entry.pack(side="left", padx=5)
        ctk.CTkLabel(r4, text="(Để hiển thị tên HS khi preview)",
                     font=("Arial", 11), text_color="#999").pack(side="left", padx=5)

        # Cột mức/điểm
        r5 = ctk.CTkFrame(sec2, fg_color="transparent")
        r5.pack(fill="x", padx=15, pady=(4,12))
        ctk.CTkLabel(r5, text="Cột Mức/Điểm:", font=("Arial", 12),
                     text_color="#333", width=130, anchor="w").pack(side="left")
        self.grade_entry = ctk.CTkEntry(r5, width=80, height=34, font=("Arial", 13),
                                         placeholder_text="VD: Q", justify="center")
        self.grade_entry.pack(side="left", padx=5)
        ctk.CTkLabel(r5, text="(Để tham chiếu xếp loại tự động)",
                     font=("Arial", 11), text_color="#999").pack(side="left", padx=5)

        # === Preview ===
        self.preview_frame = ctk.CTkFrame(body, fg_color="#FFF8F0", corner_radius=8,
                                          border_width=1, border_color="#FDEBD0")
        self.preview_frame.pack(fill="x", pady=(0,10))
        self.preview_label = ctk.CTkLabel(self.preview_frame,
                                          text="💡 Nhập thông tin rồi bấm 'Xem trước' để kiểm tra",
                                          font=("Arial", 12), text_color="#888",
                                          wraplength=450, justify="left")
        self.preview_label.pack(padx=15, pady=12)

        # === Buttons ===
        btn_frame = ctk.CTkFrame(body, fg_color="transparent")
        btn_frame.pack(fill="x", pady=(5,5))

        ctk.CTkButton(btn_frame, text="👁 Xem trước", width=120, height=36,
                      fg_color="#3498DB", hover_color="#2980B9",
                      font=("Arial", 12, "bold"),
                      command=self._preview).pack(side="left", padx=(0,10))

        ctk.CTkButton(btn_frame, text="✅ Áp dụng", width=120, height=36,
                      fg_color="#27AE60", hover_color="#219653",
                      font=("Arial", 12, "bold"),
                      command=self._apply).pack(side="left", padx=(0,10))

        ctk.CTkButton(btn_frame, text="🔄 Reset", width=100, height=36,
                      fg_color="#E74C3C", hover_color="#C0392B",
                      font=("Arial", 12),
                      command=self._reset).pack(side="left", padx=(0,10))

        ctk.CTkButton(btn_frame, text="Đóng", width=80, height=36,
                      fg_color="#95A5A6", hover_color="#7F8C8D",
                      font=("Arial", 12),
                      command=self.destroy).pack(side="right")

        # Pre-fill nếu đã có config
        if self.config.enabled:
            self.col_entry.insert(0, self.config.comment_col_letter)
            self.start_entry.insert(0, str(self.config.row_start or ""))
            self.end_entry.insert(0, str(self.config.row_end or ""))
            if self.config.name_col_letter:
                self.name_entry.insert(0, self.config.name_col_letter)
            if self.config.grade_col_letter:
                self.grade_entry.insert(0, self.config.grade_col_letter)

    def _validate(self):
        """Validate input, trả về True nếu hợp lệ"""
        col_str = self.col_entry.get().strip()
        start_str = self.start_entry.get().strip()
        end_str = self.end_entry.get().strip()

        if not col_str:
            messagebox.showwarning("Thiếu thông tin", "Vui lòng nhập CỘT nhận xét!", parent=self)
            return False
        if not start_str:
            messagebox.showwarning("Thiếu thông tin", "Vui lòng nhập DÒNG bắt đầu!", parent=self)
            return False
        if not end_str:
            messagebox.showwarning("Thiếu thông tin", "Vui lòng nhập DÒNG kết thúc!", parent=self)
            return False

        col_num = col_letter_to_number(col_str)
        if col_num is None or col_num < 1:
            messagebox.showerror("Sai dữ liệu", f"Cột '{col_str}' không hợp lệ!\nNhập A, B, S, AB... hoặc số.", parent=self)
            return False

        try:
            start = int(start_str)
            end = int(end_str)
        except ValueError:
            messagebox.showerror("Sai dữ liệu", "Dòng bắt đầu/kết thúc phải là số nguyên!", parent=self)
            return False

        if start < 1 or end < 1:
            messagebox.showerror("Sai dữ liệu", "Dòng phải >= 1!", parent=self)
            return False
        if end < start:
            messagebox.showerror("Sai dữ liệu", f"Dòng kết thúc ({end}) phải >= dòng bắt đầu ({start})!", parent=self)
            return False

        return True

    def _parse_config(self):
        """Parse input thành config object"""
        col_str = self.col_entry.get().strip().upper()
        col_num = col_letter_to_number(col_str)
        col_letter = get_column_letter(col_num) if col_num else col_str

        self.config.enabled = True
        self.config.comment_col = col_num
        self.config.comment_col_letter = col_letter
        self.config.row_start = int(self.start_entry.get().strip())
        self.config.row_end = int(self.end_entry.get().strip())

        # Tùy chọn
        name_str = self.name_entry.get().strip()
        if name_str:
            name_num = col_letter_to_number(name_str)
            if name_num:
                self.config.name_col = name_num
                self.config.name_col_letter = get_column_letter(name_num)

        grade_str = self.grade_entry.get().strip()
        if grade_str:
            grade_num = col_letter_to_number(grade_str)
            if grade_num:
                self.config.grade_col = grade_num
                self.config.grade_col_letter = get_column_letter(grade_num)

    def _preview(self):
        """Hiển thị preview dữ liệu sẽ được điền"""
        if not self._validate():
            return
        self._parse_config()

        total = self.config.row_end - self.config.row_start + 1
        preview_text = f"✅ SẼ ĐIỀN VÀO:\n"
        preview_text += f"  Cột: {self.config.comment_col_letter} (cột số {self.config.comment_col})\n"
        preview_text += f"  Từ dòng {self.config.row_start} → dòng {self.config.row_end}\n"
        preview_text += f"  Tổng: {total} ô cần điền\n"

        # Preview từ worksheet nếu có
        if self.ws:
            preview_text += f"\n📋 DỮ LIỆU MẪU (5 dòng đầu):\n"
            for r in range(self.config.row_start, min(self.config.row_start + 5, self.config.row_end + 1)):
                name = ""
                if self.config.name_col:
                    name_val = self.ws.cell(r, self.config.name_col).value
                    name = str(name_val)[:20] if name_val else "(trống)"
                existing = self.ws.cell(r, self.config.comment_col).value
                status = f'"{existing[:30]}..."' if existing else "(trống — sẽ điền)"
                if name:
                    preview_text += f"  {self.config.comment_col_letter}{r}: {name} → {status}\n"
                else:
                    preview_text += f"  {self.config.comment_col_letter}{r}: {status}\n"

        if self.config.name_col:
            preview_text += f"\n👤 Cột Họ tên: {self.config.name_col_letter}"
        if self.config.grade_col:
            preview_text += f"\n📊 Cột Mức/Điểm: {self.config.grade_col_letter}"

        self.preview_label.configure(text=preview_text, text_color="#333")

    def _apply(self):
        """Áp dụng cấu hình"""
        if not self._validate():
            return
        self._parse_config()

        total = self.config.row_end - self.config.row_start + 1
        msg = (f"Áp dụng cấu hình thủ công:\n\n"
               f"• Cột nhận xét: {self.config.comment_col_letter}\n"
               f"• Dòng: {self.config.row_start} → {self.config.row_end}\n"
               f"• Tổng: {total} ô\n\n"
               f"Bấm OK để xác nhận.")

        if messagebox.askyesno("Xác nhận", msg, parent=self):
            if self.on_apply:
                self.on_apply(self.config)
            self.destroy()

    def _reset(self):
        """Reset về auto-detect"""
        self.config.reset()
        self.col_entry.delete(0, "end")
        self.start_entry.delete(0, "end")
        self.end_entry.delete(0, "end")
        self.name_entry.delete(0, "end")
        self.grade_entry.delete(0, "end")
        self.preview_label.configure(
            text="🔄 Đã reset! App sẽ dùng chế độ tự nhận diện.",
            text_color="#E67E22")
        if self.on_apply:
            self.on_apply(self.config)
