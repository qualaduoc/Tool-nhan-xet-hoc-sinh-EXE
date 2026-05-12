# excel_processor.py - Xử lý đọc/ghi file Excel
import openpyxl
from openpyxl.utils import get_column_letter
import os
import re

class ExcelProcessor:
    """Đọc file Excel giáo viên tải lên, nhận diện cấu trúc, điền nhận xét"""

    # Mapping mức đánh giá - dùng LOWERCASE để so khớp linh hoạt
    # Tiểu học: T (Tốt/Hoàn thành tốt), H (Hoàn thành/Đạt), C (Chưa hoàn thành)
    LEVEL_MAP_TIEU_HOC = {
        # Viết tắt 1 ký tự
        "t": "T", "h": "H", "c": "C", "đ": "H", "d": "H",
        # Viết tắt nhiều ký tự
        "htt": "T", "ht": "H", "cht": "C", "kht": "C",
        # Viết đầy đủ
        "tốt": "T", "hoàn thành tốt": "T",
        "hoàn thành": "H", "đạt": "H",
        "chưa hoàn thành": "C", "không hoàn thành": "C", "chưa đạt": "C",
        # Biến thể khác
        "t ": "T", "đ ": "H", "h ": "H", "c ": "C",  # có dấu cách thừa
    }

    # THCS: XS, T, K, D, CD
    LEVEL_MAP_THCS = {
        # Viết tắt
        "xs": "XS", "t": "T", "g": "T", "k": "K", "đ": "D", "d": "D",
        "tb": "D", "cd": "CD", "cđ": "CD", "y": "CD",
        # Viết đầy đủ
        "xuất sắc": "XS", "xuat sac": "XS",
        "tốt": "T", "giỏi": "T",
        "khá": "K", "kha": "K",
        "đạt": "D", "dat": "D", "trung bình": "D", "trung binh": "D",
        "chưa đạt": "CD", "chua dat": "CD", "yếu": "CD", "yeu": "CD", "kém": "CD", "kem": "CD",
        # Hoàn thành mapping (Tiểu học format nhưng trong file THCS)
        "hoàn thành tốt": "T", "htt": "T",
        "hoàn thành": "K", "ht": "K",
        "chưa hoàn thành": "CD", "cht": "CD", "kht": "CD", "không hoàn thành": "CD",
    }

    # THPT (TT22): T, K, Đ, CĐ
    LEVEL_MAP_THPT = {
        # Viết tắt
        "t": "T", "g": "T", "k": "K", "đ": "D", "d": "D",
        "tb": "D", "cd": "CD", "cđ": "CD", "y": "CD",
        # Viết đầy đủ
        "tốt": "T", "giỏi": "T", "xuất sắc": "T",
        "khá": "K", "kha": "K",
        "đạt": "D", "dat": "D", "trung bình": "D", "trung binh": "D",
        "chưa đạt": "CD", "chua dat": "CD", "yếu": "CD", "yeu": "CD", "kém": "CD", "kem": "CD",
    }

    def __init__(self):
        self.wb = None
        self.filepath = None
        self.file_type = None  # "nlpc" or "dinhky_monhoc"

    def load_file(self, filepath):
        """Tải file Excel — hỗ trợ cả .xlsx và .xls"""
        self.filepath = filepath
        ext = os.path.splitext(filepath)[1].lower()

        if ext == ".xls":
            self.wb = self._convert_xls_to_openpyxl(filepath)
        else:
            self.wb = openpyxl.load_workbook(filepath)

        self.file_type = self._detect_file_type()
        return self.file_type

    def _convert_xls_to_openpyxl(self, filepath):
        """Chuyển file .xls (Excel 97-2003) sang openpyxl workbook"""
        import xlrd
        xls_wb = xlrd.open_workbook(filepath)
        new_wb = openpyxl.Workbook()
        # Xóa sheet mặc định
        new_wb.remove(new_wb.active)

        for sn in xls_wb.sheet_names():
            xls_ws = xls_wb.sheet_by_name(sn)
            new_ws = new_wb.create_sheet(title=sn)
            for r in range(xls_ws.nrows):
                for c in range(xls_ws.ncols):
                    cell = xls_ws.cell(r, c)
                    val = cell.value
                    # xlrd trả date dạng float, convert sang string
                    if cell.ctype == xlrd.XL_CELL_DATE:
                        try:
                            dt = xlrd.xldate_as_datetime(val, xls_wb.datemode)
                            val = dt.strftime("%d/%m/%Y")
                        except Exception:
                            pass
                    elif cell.ctype == xlrd.XL_CELL_NUMBER:
                        # Nếu số nguyên, bỏ .0
                        if val == int(val):
                            val = int(val)
                    new_ws.cell(row=r + 1, column=c + 1, value=val)
        return new_wb

    def _detect_file_type(self):
        """Nhận diện loại file: NLPC, đánh giá theo môn, hoặc đánh giá định kỳ môn học"""
        if not self.wb:
            return None

        first_sheet = self.wb[self.wb.sheetnames[0]]

        # Scan 10 row đầu để nhận diện
        for row in first_sheet.iter_rows(min_row=1, max_row=min(10, first_sheet.max_row), values_only=False):
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    val = cell.value.strip().lower()
                    # NLPC format
                    if "năng lực" in val or "phẩm chất" in val:
                        return "nlpc"
                    # Đánh giá định kỳ môn học (có "Mức đạt được" + "Nội dung" hoặc "Nhận xét")
                    if "mức đạt" in val:
                        return "dinhky_monhoc"
                    if "nội dung nhận xét" in val:
                        return "dinhky_monhoc"

        # Fallback: check các sheet khác
        for sn in self.wb.sheetnames:
            ws = self.wb[sn]
            for cell in ws[1]:
                if cell.value and isinstance(cell.value, str):
                    if "mức đạt được" in cell.value.lower() or "nội dung nhận xét" in cell.value.lower():
                        return "dinhky_monhoc"

        return "unknown"

    def get_sheet_info(self):
        """Trả về thông tin các sheet và số dòng dữ liệu"""
        if not self.wb:
            return []
        info = []
        for sn in self.wb.sheetnames:
            if sn.lower() == "huongdan":
                continue
            ws = self.wb[sn]
            data_rows = 0
            for row in ws.iter_rows(min_row=2, values_only=True):
                if any(v is not None for v in row):
                    data_rows += 1
            info.append({"name": sn, "rows": data_rows, "max_col": ws.max_column})
        return info

    def get_preview_data(self, sheet_name, max_rows=10):
        """Xem trước dữ liệu sheet — tự tìm header và data start"""
        if not self.wb or sheet_name not in self.wb.sheetnames:
            return [], []
        ws = self.wb[sheet_name]

        # Tìm dòng header: dòng có "STT" hoặc "Họ" hoặc "Mức đạt"
        header_row = 1
        for r in range(1, min(12, ws.max_row + 1)):
            for c in range(1, min(ws.max_column + 1, 15)):
                v = ws.cell(r, c).value
                if v and isinstance(v, str):
                    vl = v.strip().lower()
                    if vl == "stt" or "họ và tên" in vl or "họ tên" in vl:
                        header_row = r
                        break
            if header_row > 1:
                break

        # Tìm dòng data: dòng đầu tiên sau header có STT là số
        data_start = header_row + 1
        for r in range(header_row + 1, min(header_row + 5, ws.max_row + 1)):
            v = ws.cell(r, 1).value
            if v is not None:
                try:
                    int(v)
                    data_start = r
                    break
                except (ValueError, TypeError):
                    continue

        # Thu thập header: merge từ header_row đến data_start-1
        max_col = min(ws.max_column, 15)
        headers = []
        for c in range(1, max_col + 1):
            # Ưu tiên row cuối cùng trong header range (thường có tên cột chi tiết nhất)
            best = ""
            for r in range(header_row, data_start):
                v = ws.cell(r, c).value
                if v and isinstance(v, str) and len(v.strip()) > len(best):
                    best = v.strip()
            headers.append(best)

        # Thêm thông tin trường/lớp/môn vào dòng đầu preview
        info_rows = []
        for r in range(1, header_row):
            v = ws.cell(r, 1).value
            if v and isinstance(v, str) and v.strip():
                info_row = [""] * max_col
                info_row[0] = v.strip()
                info_rows.append(info_row)

        # Thu thập data
        rows_data = []
        for row in ws.iter_rows(min_row=data_start, max_row=min(ws.max_row, data_start + max_rows - 1),
                                min_col=1, max_col=max_col, values_only=True):
            rows_data.append([str(v) if v else "" for v in row])

        return headers, info_rows + rows_data

    def process_nlpc(self, comment_bank, cap="tieu_hoc"):
        """Xử lý file NLPC (Năng lực Phẩm chất) - hỗ trợ Tiểu học + THCS"""
        ws = self.wb[self.wb.sheetnames[0]]
        count = 0

        # Tìm vị trí cột từ header R1-R2
        nlc_cols = []     # NL chung: mức T/Đ/C
        nldt_cols = []    # NL đặc thù: mức T/Đ/C
        pc_cols = []      # Phẩm chất: mức T/Đ/C
        nx_nlc_col = None   # Nội dung nhận xét NLC
        nx_nldt_col = None  # Nội dung nhận xét NLĐT
        nx_pc_col = None    # Nội dung nhận xét PC

        # Scan header R1 để tìm nhóm cột
        current_group = None
        for c in range(1, min(ws.max_column + 1, 30)):
            v1 = ws.cell(1, c).value
            v2 = ws.cell(2, c).value
            h1 = str(v1).strip().lower() if v1 else ""
            h2 = str(v2).strip().lower() if v2 else ""

            # Xác định nhóm từ R1
            if "năng lực chung" in h1:
                current_group = "nlc"
            elif "năng lực đặc thù" in h1:
                current_group = "nldt"
            elif "phẩm chất" in h1:
                current_group = "pc"
            elif "nhận xét năng lực chung" in h1:
                current_group = None
                # Tìm cột "Nội dung" trong nhóm NX NLC
                if h2 and "nội dung" in h2:
                    nx_nlc_col = c
                elif ws.cell(2, c + 1).value and "nội dung" in str(ws.cell(2, c + 1).value).lower():
                    nx_nlc_col = c + 1
            elif "nhận xét năng lực đặc thù" in h1:
                current_group = None
                if h2 and "nội dung" in h2:
                    nx_nldt_col = c
                elif ws.cell(2, c + 1).value and "nội dung" in str(ws.cell(2, c + 1).value).lower():
                    nx_nldt_col = c + 1
            elif "nhận xét phẩm chất" in h1:
                current_group = None
                if h2 and "nội dung" in h2:
                    nx_pc_col = c
                elif ws.cell(2, c + 1).value and "nội dung" in str(ws.cell(2, c + 1).value).lower():
                    nx_pc_col = c + 1

            # Thu thập cột mức vào nhóm
            if current_group and h2 and h2 not in ("mã nhận xét", "nội dung"):
                if current_group == "nlc":
                    nlc_cols.append(c)
                elif current_group == "nldt":
                    nldt_cols.append(c)
                elif current_group == "pc":
                    pc_cols.append(c)

        # Fallback nếu không tìm thấy header: dùng hardcode cũ
        if not nlc_cols:
            nlc_cols = list(range(5, 8))    # E-G
        if not nldt_cols:
            nldt_cols = list(range(8, 15))  # H-N
        if not pc_cols:
            pc_cols = list(range(15, 20))   # O-S
        if not nx_nlc_col:
            nx_nlc_col = 21   # U
        if not nx_nldt_col:
            nx_nldt_col = 23  # W
        if not nx_pc_col:
            nx_pc_col = 25    # Y

        # Xử lý từng học sinh
        for row_idx in range(3, ws.max_row + 1):
            name_cell = ws.cell(row=row_idx, column=3).value
            if not name_cell:
                continue

            # Tính mức riêng cho từng nhóm (NLPC dùng T/Đ/C, kho dùng T/D/C)
            NLPC_MAP = {"t": "T", "đ": "D", "d": "D", "c": "C",
                        "tốt": "T", "đạt": "D", "chưa đạt": "C"}

            def get_group_level(cols):
                levels = []
                for c in cols:
                    val = ws.cell(row=row_idx, column=c).value
                    if val and isinstance(val, str):
                        mapped = NLPC_MAP.get(val.strip().lower(), val.strip())
                        levels.append(mapped)
                if not levels:
                    return "D"
                from collections import Counter
                counter = Counter(levels)
                return counter.most_common(1)[0][0]

            nlc_level = get_group_level(nlc_cols)
            nldt_level = get_group_level(nldt_cols)
            pc_level = get_group_level(pc_cols)

            # Điền nhận xét NLC
            if not ws.cell(row=row_idx, column=nx_nlc_col).value:
                comment = comment_bank.get_random_comment(cap, "nlpc", "nang_luc_chung", nlc_level)
                if comment:
                    ws.cell(row=row_idx, column=nx_nlc_col).value = comment

            # Điền nhận xét NLĐT
            if not ws.cell(row=row_idx, column=nx_nldt_col).value:
                comment = comment_bank.get_random_comment(cap, "nlpc", "nang_luc_dac_thu", nldt_level)
                if comment:
                    ws.cell(row=row_idx, column=nx_nldt_col).value = comment

            # Điền nhận xét PC
            if not ws.cell(row=row_idx, column=nx_pc_col).value:
                comment = comment_bank.get_random_comment(cap, "nlpc", "pham_chat", pc_level)
                if comment:
                    ws.cell(row=row_idx, column=nx_pc_col).value = comment

            count += 1
        return count

    def process_monhoc(self, comment_bank, cap="tieu_hoc"):
        """Xử lý file đánh giá theo môn học (cả định kỳ và thường xuyên)"""
        count = 0
        for sn in self.wb.sheetnames:
            if sn.lower() == "huongdan":
                continue
            ws = self.wb[sn]

            # Tìm cặp (Mức đạt được, Nội dung nhận xét) — scan 10 row đầu
            pairs = self._find_muc_nhanxet_pairs(ws)
            if not pairs:
                continue

            # Tìm tên môn học từ header (R1-R6)
            subject_name = self._detect_subject_name(ws, sn)

            # Tìm dòng bắt đầu data
            data_start = self._find_data_start(ws)

            for muc_col, nhanxet_col in pairs:
                for row_idx in range(data_start, ws.max_row + 1):
                    muc_val = ws.cell(row=row_idx, column=muc_col).value
                    existing = ws.cell(row=row_idx, column=nhanxet_col).value

                    if not muc_val or existing:
                        continue

                    level = str(muc_val).strip()
                    normalized = self._normalize_level(level, cap)

                    comment = comment_bank.get_random_comment(cap, "mon_hoc", subject_name, normalized)
                    if not comment:
                        comment = self._fallback_comment(comment_bank, cap, subject_name, normalized)

                    if comment:
                        ws.cell(row=row_idx, column=nhanxet_col).value = comment
                        count += 1
        return count

    def _find_muc_nhanxet_pairs(self, ws):
        """Tìm tất cả cặp (cột Mức đạt, cột Nội dung) trong header"""
        pairs = []
        # Scan 10 row đầu để tìm
        muc_cols = []
        noidung_cols = []
        for r in range(1, min(11, ws.max_row + 1)):
            for c in range(1, min(ws.max_column + 1, 50)):
                v = ws.cell(r, c).value
                if v and isinstance(v, str):
                    vl = v.strip().lower()
                    if "mức đạt" in vl:
                        muc_cols.append(c)
                    elif "nội dung" in vl and ("nhận xét" in vl or r >= 8):
                        noidung_cols.append(c)
                    # Fallback: cột chỉ ghi "Nội dung" (row 9 của file đánh giá định kỳ)
                    elif vl == "nội dung":
                        noidung_cols.append(c)

        # Ghép cặp: mỗi Mức đạt → tìm Nội dung gần nhất bên phải
        for mc in muc_cols:
            best_nd = None
            best_dist = 999
            for nd in noidung_cols:
                if nd > mc and (nd - mc) < best_dist:
                    best_nd = nd
                    best_dist = nd - mc
            if best_nd:
                pairs.append((mc, best_nd))

        # Fallback: format cũ (row 1 có "mức đạt" + "nội dung nhận xét")
        if not pairs:
            muc_col = None
            nhanxet_col = None
            for cell in ws[1]:
                if cell.value and isinstance(cell.value, str):
                    val = cell.value.strip().lower()
                    if "mức đạt" in val:
                        muc_col = cell.column
                    elif "nội dung nhận xét" in val or "nội dung" in val:
                        nhanxet_col = cell.column
            if muc_col and nhanxet_col:
                pairs.append((muc_col, nhanxet_col))

        return pairs

    def _detect_subject_name(self, ws, sheet_name):
        """Tìm tên môn học từ header file"""
        import re
        # Tìm trong R1-R6: dòng chứa 'ĐÁNH GIÁ' hoặc 'MÔN'
        for r in range(1, 7):
            v = ws.cell(r, 1).value
            if v and isinstance(v, str):
                vl = v.strip()
                vu = vl.upper()
                if "ĐÁNH GIÁ" in vu and "MÔN" in vu:
                    # Ưu tiên tên trong ngoặc: "TH-CN (Tin Học)" → "Tin Học"
                    m_paren = re.search(r'\(([^)]+)\)', vl)
                    if m_paren:
                        return m_paren.group(1).strip()
                    # Fallback: lấy sau "MÔN HỌC" hoặc "MÔN"
                    m = re.search(r'MÔN\s+(?:HỌC\s+)?(.+)', vu)
                    if m:
                        return m.group(1).strip().title()
        # Fallback: dùng tên sheet
        return sheet_name.strip()

    def _find_data_start(self, ws):
        """Tìm dòng đầu tiên có data học sinh (có STT số ở cột A)"""
        for r in range(2, min(15, ws.max_row + 1)):
            v = ws.cell(r, 1).value
            if v is not None:
                try:
                    int(v)
                    return r
                except (ValueError, TypeError):
                    continue
        return 2  # Fallback

    def _fallback_comment(self, comment_bank, cap, subject_name, level):
        """Tìm nhận xét từ môn học có tên gần giống"""
        subjects = comment_bank.data.get(cap, {}).get("mon_hoc", {})
        sn_lower = subject_name.lower()
        for key in subjects:
            if key.lower() in sn_lower or sn_lower in key.lower():
                comments = comment_bank.get_comments(cap, "mon_hoc", key, level)
                if comments:
                    import random
                    return random.choice(comments)

        # Nếu vẫn không có, dùng nhận xét mức chung (cho tất cả cấp)
        muc_chung = comment_bank.data.get(cap, {}).get("muc_chung", {}).get(level, {})
        if isinstance(muc_chung, dict) and "nhan_xet" in muc_chung:
            pool = muc_chung["nhan_xet"]
            if pool:
                import random
                return random.choice(pool)

        # Fallback cuối: lấy nhận xét từ bất kỳ môn nào có cùng mức
        for key in subjects:
            comments = comment_bank.get_comments(cap, "mon_hoc", key, level)
            if comments:
                import random
                return random.choice(comments)

        return ""

    def _normalize_level(self, level, cap):
        """Chuẩn hóa mức đánh giá - xử lý mọi biến thể viết tắt/đầy đủ"""
        raw = level.strip()
        key = raw.lower().strip()

        if cap == "tieu_hoc":
            level_map = self.LEVEL_MAP_TIEU_HOC
            default = "H"
        elif cap == "thpt":
            level_map = self.LEVEL_MAP_THPT
            default = "D"
        else:
            level_map = self.LEVEL_MAP_THCS
            default = "D"

        # Thử exact match (lowercase)
        if key in level_map:
            return level_map[key]

        # Thử với dấu cách thừa ở cuối (phổ biến trong file Excel)
        key_stripped = key.rstrip()
        if key_stripped in level_map:
            return level_map[key_stripped]

        # Thử match từ dài nhất trước (ưu tiên "hoàn thành tốt" trước "hoàn thành")
        sorted_keys = sorted(level_map.keys(), key=len, reverse=True)
        for k in sorted_keys:
            if k in key:
                return level_map[k]

        return default

    def _get_majority_level(self, levels, cap):
        """Xác định mức đánh giá chung - ưu tiên mức thấp nhất nếu có"""
        if not levels:
            return "T" if cap == "tieu_hoc" else "D"

        normalized = [self._normalize_level(l, cap) for l in levels]
        from collections import Counter
        counter = Counter(normalized)

        # Nếu tất cả cùng mức → trả về mức đó
        if len(counter) == 1:
            return counter.most_common(1)[0][0]

        # Nếu có nhiều mức → dùng mức chiếm đa số, nhưng nếu có "C" thì ưu tiên hạ
        if cap == "tieu_hoc":
            priority = {"C": 0, "H": 1, "T": 2}
        else:
            priority = {"CD": 0, "D": 1, "K": 2, "T": 3, "XS": 4}

        # Nếu mức thấp chiếm >= 30% → dùng mức thấp
        total = sum(counter.values())
        for muc in sorted(counter.keys(), key=lambda x: priority.get(x, 99)):
            if counter[muc] / total >= 0.3:
                return muc

        return counter.most_common(1)[0][0]

    def save_output(self, output_path):
        """Lưu file kết quả"""
        if self.wb:
            self.wb.save(output_path)
            return True
        return False
